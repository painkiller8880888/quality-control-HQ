import sys
from contextlib import contextmanager
import os
from types import ModuleType, SimpleNamespace
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import MagicMock, patch

from django.test import TestCase, override_settings
from django.core.files.uploadedfile import SimpleUploadedFile
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIClient

from .models import (
    AppSetting,
    ClassMaster,
    History,
    InspectionFile,
    InspectionSession,
    InspectionTarget,
    Job,
    LayoutMaster,
    LayoutObject,
    Machine,
    MachineAssignment,
    Master,
    MasterClass,
    Structure,
    User,
)


def authenticate_test_client(client, login_name="test-admin", role=User.Role.ADMIN):
    user = User.objects.create(login_name=login_name, display_name=login_name, password_hash="!", role=role)
    client.force_authenticate(user)
    return user


def make_csv_row(root_code, parent_code, child_code, child_name, level, department, quantity):
    cols = [""] * 36
    cols[1] = root_code
    cols[8] = parent_code
    cols[9] = child_code
    cols[10] = child_name
    cols[11] = str(level)
    cols[16] = department
    cols[18] = str(quantity)
    return ",".join(f'"{c}"' for c in cols)


CSV_HEADER = ",".join(f'"h{i}"' for i in range(36))

MASTER_CSV_LINES = [
    # 0: root BA0061 (親は自分自身)
    make_csv_row("BA0061", "BA0061", "BA0061", "DYL完成品", 1, "生産管理部", 0),
    # 1: BA0061 の子 CAG0008 (鍍金)
    make_csv_row("BA0061", "BA0061", "CAG0008", "DYL鍍金 NI", 2, "IFC-カクテル", 0),
    # 2: CAG0008 の子 CAL0001 (加工タッピング)
    make_csv_row("BA0061", "CAG0008", "CAL0001", "DYLタッピング", 3, "生産管理部", 0),
    # 3: CAL0001 の子 DA0045 (ダイカスト)
    make_csv_row("BA0061", "CAL0001", "DA0045", "DYL鋳物", 4, "IFC-特殊工程", 0),
    # 4: CAG0008 の子 DC0034 (ネジ) - 同一親に複数子
    make_csv_row("BA0061", "CAG0008", "DC0034", "ビス 5X10", 3, "IFC-ネジ専門", 0),
    # 5: 異なるルート BA0099
    make_csv_row("BA0099", "BA0099", "BA0099", "DCYK-32完成品", 1, "生産管理部", 0),
    # 6: BA0099 の子 CAG0010 (鍍金)
    make_csv_row("BA0099", "BA0099", "CAG0010", "DCYK鍍金 NI", 2, "IFC-カクテル", 0),
    # 7: プレスのコード DK0289 (node_type_1=プレスになる)
    make_csv_row("BA0080", "CAD0001", "DK0289", "DC100プレス部品", 5, "IFC-本社プレス工場", 1.000),
]

MASTER_CSV = (CSV_HEADER + "\n" + "\n".join(MASTER_CSV_LINES)).encode("cp932")


@contextmanager
def mock_excel_com_modules():
    """Provide only the COM module surface used by mocked Excel tests on Linux."""
    if sys.platform == "win32":
        import pythoncom
        import win32com.client as win32_client

        yield pythoncom, win32_client
        return

    class ComError(Exception):
        pass

    pythoncom = ModuleType("pythoncom")
    pythoncom.com_error = ComError
    pythoncom.CoInitialize = MagicMock()
    pythoncom.CoUninitialize = MagicMock()

    win32com = ModuleType("win32com")
    win32_client = ModuleType("win32com.client")
    win32_client.GetActiveObject = MagicMock()
    win32_client.Dispatch = MagicMock()
    win32_client.DispatchEx = MagicMock()
    win32com.client = win32_client

    with patch.dict(
        sys.modules,
        {
            "pythoncom": pythoncom,
            "win32com": win32com,
            "win32com.client": win32_client,
        },
    ):
        yield pythoncom, win32_client


def csv_with_codes(codes_with_data):
    """Build a CSV string from a list of (root, parent, child, name, level, dept, qty) tuples."""
    lines = [CSV_HEADER]
    for row in codes_with_data:
        lines.append(make_csv_row(*row))
    return "\n".join(lines).encode("cp932")


class PhaseOneApiTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = authenticate_test_client(self.client)
        master1 = Master.objects.create(code="C1234", name="ホルダーAssy", category=1)
        master6 = Master.objects.create(code="C5678", name="検査品", category=6)
        InspectionFile.objects.create(master=master1, file_name="C1234.xlsx", file_path=r"C:\製品検査\(1)\C1234.xlsx")
        InspectionFile.objects.create(master=master6, file_name="C5678.xlsx", file_path=r"C:\製品検査\(1)\C5678.xlsx")

    def test_manual_targets_preserve_unknown_code_warning(self):
        response = self.client.post(
            "/api/inspection-targets/manual/",
            {"date": "2026-05-20", "codes": ["C1234", "X9999"]},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        targets = self.client.get("/api/inspection-targets/?date=2026-05-20").json()

        self.assertEqual(len(targets), 2)
        unknown = next(target for target in targets if target["code"] == "X9999")
        self.assertEqual(unknown["warnings"][0]["error_code"], "UNKNOWN_CODE")

    def test_bulk_history_upsert_creates_and_removes_checked_slots(self):
        self.client.post(
            "/api/inspection-targets/manual/",
            {"date": "2026-05-20", "codes": ["C1234"]},
            format="json",
        )
        target_id = InspectionTarget.objects.get(normalized_code="C1234").id
        response = self.client.post(
            "/api/history/bulk-upsert/",
            {
                "date": "2026-05-20",
                "items": [
                    {
                        "target_id": target_id,
                        "checks": {"A": True, "B": False, "C": True, "D": False},
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(set(History.objects.values_list("time_slot", flat=True)), {"A", "C"})

        response = self.client.patch(
            "/api/history/",
            {"date": "2026-05-20", "target_id": target_id, "time": "C", "checked": False},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            set(History.objects.filter(deleted_at__isnull=True).values_list("time_slot", flat=True)),
            {"A"},
        )
        self.assertTrue(History.objects.filter(time_slot="C", deleted_by=self.user, deleted_at__isnull=False).exists())

    def test_daily_report_generation_writes_expected_cells(self):
        with TemporaryDirectory() as temp_dir:
            template_path = Path(temp_dir) / "daily.xlsm"
            output_dir = Path(temp_dir) / "reports"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "日報"
            workbook.save(template_path)

            History.objects.create(
                date="2026-05-20",
                master=Master.objects.get(code="C1234"),
                time_slot="A",
                created_by=self.user,
            )
            History.objects.create(
                date="2026-05-20",
                master=Master.objects.get(code="C5678"),
                time_slot="B",
                created_by=self.user,
            )

            with override_settings(
                DAILY_REPORT_TEMPLATE=template_path,
                DAILY_REPORT_OUTPUT_DIR=output_dir,
            ):
                response = self.client.post(
                    "/api/daily-report/generate/",
                    {"date": "2026-05-20"},
                    format="json",
                )

            self.assertEqual(response.status_code, 202)
            job = Job.objects.get(job_id=response.json()["job_id"])
            self.assertEqual(job.status, Job.Status.SUCCEEDED)

            report = load_workbook(output_dir / f"2026-05-20_u{self.user.pk}.xlsm")
            sheet = report["日報"]
            self.assertEqual(sheet["C26"].value, "ホルダーAssy")
            self.assertEqual(sheet["F18"].value, "検査品")

    def test_plan_import_merges_ocr_text_and_excel_codes(self):
        ocr_master = Master.objects.create(code="CDP0028", name="OCR対象", category=5)
        excel_master = Master.objects.create(code="CAP0044", name="Excel対象", category=4)
        machine = Machine.objects.create(machine_no="PLAN-M", machine_name="Plan", machine_class=2, shape_type="rectangle", map_x=0, map_y=0, width=1, height=1)
        MachineAssignment.objects.create(machine=machine, code=ocr_master, assignment_class=2)
        InspectionFile.objects.create(master=ocr_master, file_name="CDP0028.xlsx", file_path=r"C:\製品検査\(2)\CDP0028.xlsx")
        InspectionFile.objects.create(master=excel_master, file_name="CAP0044.xlsx", file_path=r"C:\製品検査\(1)\CAP0044.xlsx")

        with TemporaryDirectory() as temp_dir:
            excel_path = Path(temp_dir) / "plan.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "雛形"
            worksheet["F4"] = "cap0044"
            worksheet["F5"] = "cdp0028"
            workbook.save(excel_path)

            with excel_path.open("rb") as handle:
                response = self.client.post(
                    "/api/plans/import/",
                    {
                        "target_date": "2026-05-21",
                        "scan_file": SimpleUploadedFile(
                            "ocr.txt",
                            b"cdp0028\ncdp0029\n",
                            content_type="text/plain",
                        ),
                        "excel_file": SimpleUploadedFile(
                            "plan.xlsx",
                            handle.read(),
                            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        ),
                        "sheet_name": "雛形",
                    },
                    format="multipart",
                )

        self.assertEqual(response.status_code, 202)
        targets = self.client.get("/api/inspection-targets/?date=2026-05-21").json()
        self.assertEqual([(t["code"], t["category"]) for t in targets], [("CAP0044", 6), ("CDP0028", 2), ("CDP0028", 7), ("CDP0029", None)])
        self.assertEqual(next(t for t in targets if t["code"] == "CDP0029")["warnings"][0]["error_code"], "UNKNOWN_CODE")

    def test_master_update_imports_csv_and_resolves_name(self):
        csv_content = (
            '"h0","h1","h2","h3","h4","h5","h6","h7","h8","品目コード","品目略称","h11","h12","h13","h14","h15","オーダー先略称"\n'
            '"0","0","0","0","0","0","0","0","0","CAP0048","テスト品名","0","0","0","0","0","生産管理部"\n'
        ).encode("cp932")

        response = self.client.post(
            "/api/master/update/",
            {"master_file": SimpleUploadedFile("master.csv", csv_content, content_type="text/csv")},
            format="multipart",
        )

        self.assertEqual(response.status_code, 202)
        self.assertTrue(Master.objects.filter(code="CAP0048", name="テスト品名").exists())
        imported_master = Master.objects.get(code="CAP0048")
        InspectionFile.objects.create(master=imported_master, file_name="CAP0048.xlsx", file_path=r"C:\製品検査\(1)\CAP0048.xlsx")

        self.client.post(
            "/api/inspection-targets/manual/",
            {"date": "2026-05-21", "codes": ["cap0048"]},
            format="json",
        )
        targets = self.client.get("/api/inspection-targets/?date=2026-05-21").json()
        self.assertEqual(targets[0]["name"], "テスト品名")

    @patch("quality.services.extract_codes_and_match_failures_from_pdf")
    def test_plan_import_from_pdf_tracks_match_failed_and_unknown_code(self, mocked_pdf_ocr):
        mocked_pdf_ocr.return_value = {
            "codes": ["CAP0048", "CAP0048", "ZZZ9999"],
            "match_failed_count": 2,
            "page_count": 3,
            "mode": "ocr",
        }
        master = Master.objects.create(code="CAP0048", name="OCR参照品名", category=1)
        machine = Machine.objects.create(machine_no="PDF-M", machine_name="PDF", machine_class=1, shape_type="rectangle", map_x=0, map_y=0, width=1, height=1)
        MachineAssignment.objects.create(machine=machine, code=master, assignment_class=1)

        response = self.client.post(
            "/api/plans/import/",
            {
                "target_date": "2026-05-22",
                "scan_file": SimpleUploadedFile("scan.pdf", b"%PDF-mock", content_type="application/pdf"),
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, 202)
        job = Job.objects.get(job_id=response.json()["job_id"])
        self.assertEqual(job.result["warning_summary"]["MATCH_FAILED"], 2)
        self.assertEqual(job.result["warning_summary"]["DUPLICATE_TARGET"], 1)
        self.assertEqual(job.result["warning_summary"]["UNKNOWN_CODE"], 1)

        targets = self.client.get("/api/inspection-targets/?date=2026-05-22").json()
        codes = {target["code"]: target for target in targets}
        self.assertEqual(codes["CAP0048"]["name"], "OCR参照品名")
        self.assertEqual(codes["ZZZ9999"]["warnings"][0]["error_code"], "UNKNOWN_CODE")

    def test_factory_map_returns_layout_and_target_machine_status(self):
        master = Master.objects.get(code="C1234")
        machine = Machine.objects.create(
            machine_no="M-001",
            machine_name="検査機",
            shape_type=Machine.ShapeType.RECTANGLE,
            map_x=4,
            map_y=5,
            width=6,
            height=3,
            machine_class=1,
        )
        MachineAssignment.objects.create(machine=machine, code=master, assignment_class=1)

        self.client.post(
            "/api/inspection-targets/factory-map/",
            {"date": "2026-05-23", "machine_id": machine.id, "code": "C1234"},
            format="json",
        )
        from quality.services import upsert_targets
        upsert_targets("2026-05-23", ["ZZZ9999"], "ocr", user=self.user)

        response = self.client.get("/api/factory-map/?date=2026-05-23")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["layout"]["grid_width"], 50)
        self.assertEqual(payload["machines"][0]["status"], "pending")
        self.assertEqual(payload["machines"][0]["target_codes"], ["C1234"])
        self.assertEqual(payload["warnings"], [{"code": "ZZZ9999", "error_code": "NO_MATCHING_MACHINE"}])

    def test_factory_map_layout_can_be_saved_and_reloaded(self):
        machine = Machine.objects.create(
            machine_no="M-002",
            machine_name="配置機",
            shape_type=Machine.ShapeType.RECTANGLE,
            map_x=0,
            map_y=0,
            width=1,
            height=1,
        )

        response = self.client.put(
            "/api/factory-map/layout/",
            {
                "layout_name": "default",
                "background_image_path": "/media/maps/factory.png",
                "grid_width": 40,
                "grid_height": 30,
                "objects": [
                    {
                        "type": "machine",
                        "machine_id": machine.id,
                        "object_name": "",
                        "grid_x": 2,
                        "grid_y": 3,
                        "width": 4,
                        "height": 5,
                    },
                    {
                        "type": "path",
                        "object_name": "主通路",
                        "grid_x": 0,
                        "grid_y": 8,
                        "width": 12,
                        "height": 2,
                    },
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(LayoutMaster.objects.get(layout_name="default").grid_width, 40)
        self.assertEqual(LayoutObject.objects.count(), 2)

        payload = self.client.get("/api/factory-map/layout/").json()
        self.assertEqual(payload["background_image_path"], "")
        self.assertEqual(payload["objects"][0]["machine_name"], "配置機")
        self.assertEqual(payload["objects"][1]["type"], "path")

    def test_factory_map_save_preserves_legacy_background_but_hides_it(self):
        layout, _ = LayoutMaster.objects.get_or_create(layout_name="legacy-background")
        layout.background_image_path = "/media/maps/legacy.png"
        layout.save(update_fields=["background_image_path"])
        response = self.client.put(
            f"/api/factory-map/layout/?layout_id={layout.id}",
            {
                "layout_name": layout.layout_name,
                "background_image_path": "/media/maps/replacement.png",
                "grid_width": 42,
                "grid_height": 31,
                "objects": [],
            },
            format="json",
        )
        self.assertEqual(response.status_code, 200)
        layout.refresh_from_db()
        self.assertEqual(layout.background_image_path, "/media/maps/legacy.png")
        self.assertEqual(response.json()["background_image_path"], "")

    def test_factory_map_background_upload_route_is_stopped(self):
        response = self.client.post(
            "/api/factory-map/upload-image/",
            {"image": SimpleUploadedFile("map.png", b"not-used", content_type="image/png")},
            format="multipart",
        )
        self.assertEqual(response.status_code, 404)

    def test_factory_map_layout_rejects_invalid_object(self):
        response = self.client.put(
            "/api/factory-map/layout/",
            {
                "layout_name": "default",
                "grid_width": 50,
                "grid_height": 50,
                "objects": [
                    {
                        "type": "polygon",
                        "grid_x": 0,
                        "grid_y": 0,
                        "width": 1,
                        "height": 1,
                    }
                ],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 400)


class PhaseTwoMasterUpdateTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = authenticate_test_client(self.client)

    def _upload_csv(self, csv_bytes):
        return self.client.post(
            "/api/master/update/",
            {"master_file": SimpleUploadedFile("master.csv", csv_bytes, content_type="text/csv")},
            format="multipart",
        )

    def test_csv_import_creates_master_records(self):
        response = self._upload_csv(MASTER_CSV)
        self.assertEqual(response.status_code, 202)

        codes = set(Master.objects.values_list("code", flat=True))
        self.assertIn("BA0061", codes)
        self.assertIn("CAG0008", codes)
        self.assertIn("CAL0001", codes)
        self.assertIn("DA0045", codes)
        self.assertIn("DC0034", codes)
        self.assertIn("BA0099", codes)
        self.assertIn("CAG0010", codes)
        self.assertIn("DK0289", codes)

    def test_csv_import_creates_structure_records(self):
        response = self._upload_csv(MASTER_CSV)
        self.assertEqual(response.status_code, 202)

        self.assertTrue(Structure.objects.filter(parent_code="BA0061", child_code="CAG0008").exists())
        self.assertTrue(Structure.objects.filter(parent_code="CAG0008", child_code="CAL0001").exists())
        self.assertTrue(Structure.objects.filter(parent_code="CAL0001", child_code="DA0045").exists())
        self.assertTrue(Structure.objects.filter(parent_code="CAG0008", child_code="DC0034").exists())
        self.assertTrue(Structure.objects.filter(parent_code="BA0099", child_code="CAG0010").exists())
        self.assertTrue(Structure.objects.filter(parent_code="CAD0001", child_code="DK0289").exists())

        # Verify structure details
        s = Structure.objects.get(parent_code="CAL0001", child_code="DA0045")
        self.assertEqual(s.root_code, "BA0061")
        self.assertEqual(s.level, 4)
        self.assertIsNone(s.quantity)

        s2 = Structure.objects.get(parent_code="CAD0001", child_code="DK0289")
        self.assertEqual(float(s2.quantity), 1.000)

    def test_csv_import_creates_master_class_records(self):
        response = self._upload_csv(MASTER_CSV)
        self.assertEqual(response.status_code, 202)

        class_vals = {
            mc.master.code: mc.class_master.class_no
            for mc in MasterClass.objects.select_related("master", "class_master").all()
        }
        # 判定不能な品目をclass 8へ自動フォールバックしない。
        self.assertEqual(class_vals, {"DK0289": 4})

    def test_csv_import_applies_product_classification(self):
        response = self._upload_csv(MASTER_CSV)
        self.assertEqual(response.status_code, 202)

        ba0061 = Master.objects.get(code="BA0061")
        # BA -> product_category = "スライド丁番", node_type_1 = "", node_type_2 = ""
        self.assertEqual(ba0061.product_category, "スライド丁番")
        self.assertIsNone(ba0061.node_type_1)
        self.assertIsNone(ba0061.node_type_2)

        cag0008 = Master.objects.get(code="CAG0008")
        # CAG -> 鍍金, スライド丁番
        self.assertEqual(cag0008.node_type_1, "鍍金")
        self.assertEqual(cag0008.product_category, "スライド丁番")

        cal0001 = Master.objects.get(code="CAL0001")
        # CAL -> 加工, タッピング, スライド丁番
        self.assertEqual(cal0001.node_type_1, "加工")
        self.assertEqual(cal0001.node_type_2, "タッピング")
        self.assertEqual(cal0001.product_category, "スライド丁番")

        da0045 = Master.objects.get(code="DA0045")
        # DA -> ダイカスト
        self.assertEqual(da0045.node_type_1, "ダイカスト")
        self.assertIsNone(da0045.product_category)

        dc0034 = Master.objects.get(code="DC0034")
        # DC -> ネジ
        self.assertEqual(dc0034.node_type_1, "ネジ")

        dk0289 = Master.objects.get(code="DK0289")
        # DK -> プレス
        self.assertEqual(dk0289.node_type_1, "プレス")

    def test_duplicate_code_in_csv_only_creates_one_master(self):
        csv_lines = [CSV_HEADER]
        csv_lines.append(make_csv_row("BA0061", "BA0061", "BA0061", "First", 1, "部署A", 0))
        csv_lines.append(make_csv_row("BA0061", "BA0061", "BA0061", "Duplicate", 1, "部署B", 0))
        csv_bytes = "\n".join(csv_lines).encode("cp932")

        response = self._upload_csv(csv_bytes)
        self.assertEqual(response.status_code, 202)

        self.assertEqual(Master.objects.filter(code="BA0061").count(), 1)
        # 最後の行の値で上書きされる
        master = Master.objects.get(code="BA0061")
        self.assertEqual(master.name, "First")

    def test_structure_unique_constraint(self):
        csv_lines = [CSV_HEADER]
        csv_lines.append(make_csv_row("BA0061", "BA0061", "CAG0008", "名A", 2, "部署", 0))
        csv_lines.append(make_csv_row("BA0061", "BA0061", "CAG0008", "名B", 2, "部署", 0))
        csv_bytes = "\n".join(csv_lines).encode("cp932")

        response = self._upload_csv(csv_bytes)
        self.assertEqual(response.status_code, 202)

        self.assertEqual(Structure.objects.filter(parent_code="BA0061", child_code="CAG0008").count(), 1)

    def test_inspection_file_scanning_and_registration(self):
        with TemporaryDirectory() as temp_dir:
            folder = Path(temp_dir) / "inspection_files"
            folder.mkdir()
            (folder / "CAG0008_検査書.pdf").write_text("dummy")
            (folder / "CAL0001_検査書_ver2.pdf").write_text("dummy")
            (folder / "DA0045_図面.pdf").write_text("dummy")
            # 該当コードなし
            (folder / "other_file.txt").write_text("dummy")

            AppSetting.objects.create(
                csv_path="",
                inspection_folder_paths=[str(folder)],
                inspection_folder_priorities={str(folder): 25},
            )

            response = self._upload_csv(MASTER_CSV)
            self.assertEqual(response.status_code, 202)

            files = InspectionFile.objects.all()
            file_names = {f.file_name for f in files}
            self.assertIn("CAG0008_検査書.pdf", file_names)
            self.assertIn("CAL0001_検査書_ver2.pdf", file_names)
            self.assertIn("DA0045_図面.pdf", file_names)
            self.assertNotIn("other_file.txt", file_names)
            self.assertTrue(all(f.priority == 25 for f in files))

    def test_scanner_records_windows_creation_time_and_continues_when_unavailable(self):
        from quality.services import _file_created_at, scan_and_classify_files

        with TemporaryDirectory() as temp_dir:
            folder = Path(temp_dir)
            first = folder / "CAG0008_first.pdf"
            second = folder / "CAG0008_second.pdf"
            first.touch()
            second.touch()

            with patch("quality.services.os.name", "nt"), patch.object(
                Path, "stat", return_value=SimpleNamespace(st_ctime=1_700_000_000)
            ):
                created = _file_created_at(first)
            self.assertTrue(timezone.is_aware(created))

            with patch("quality.services._file_created_at", side_effect=lambda path: created if path == first else None):
                scanned, warnings = scan_and_classify_files([str(folder)], {str(folder): 7})

            self.assertEqual(warnings, [])
            self.assertEqual(len(scanned["CAG0008"]), 2)
            by_name = {row["file_name"]: row for row in scanned["CAG0008"]}
            self.assertEqual(by_name[first.name]["priority"], 7)
            self.assertEqual(by_name[first.name]["file_created"], created)
            self.assertIsNone(by_name[second.name]["file_created"])

    def test_inspection_file_cleared_on_reimport(self):
        with TemporaryDirectory() as temp_dir:
            folder = Path(temp_dir) / "inspection_files"
            folder.mkdir()
            (folder / "CAG0008_検査書.pdf").write_text("dummy")

            AppSetting.objects.create(
                csv_path="",
                inspection_folder_paths=[str(folder)],
            )

            self._upload_csv(MASTER_CSV)
            first_count = InspectionFile.objects.count()
            self.assertGreater(first_count, 0)

            # 再import時に古いファイルは削除される
            self._upload_csv(MASTER_CSV)
            self.assertEqual(InspectionFile.objects.count(), first_count)

    def test_master_update_with_settings_csv_path(self):
        with TemporaryDirectory() as temp_dir:
            csv_path = Path(temp_dir) / "settings_master.csv"
            csv_path.write_bytes(MASTER_CSV)

            AppSetting.objects.create(
                csv_path=str(csv_path),
                inspection_folder_paths=[],
            )

            response = self.client.post(
                "/api/master/update/",
                {"force": False},
                format="json",
            )
            self.assertEqual(response.status_code, 202)

            self.assertTrue(Master.objects.filter(code="BA0061").exists())
            self.assertTrue(Master.objects.filter(code="CAG0008").exists())

    def test_master_update_prefers_uploaded_file_over_settings(self):
        with TemporaryDirectory() as temp_dir:
            csv_path = Path(temp_dir) / "settings_master.csv"
            csv_path.write_text("dummy,data\n")

            AppSetting.objects.create(
                csv_path=str(csv_path),
                inspection_folder_paths=[],
            )

            csv_lines = [CSV_HEADER, make_csv_row("BA0061", "BA0061", "BA0061", "UploadedOnly", 1, "部署", 0)]
            csv_bytes = "\n".join(csv_lines).encode("cp932")

            response = self.client.post(
                "/api/master/update/",
                {"master_file": SimpleUploadedFile("upload.csv", csv_bytes, content_type="text/csv")},
                format="multipart",
            )
            self.assertEqual(response.status_code, 202)

            master = Master.objects.get(code="BA0061")
            self.assertEqual(master.name, "UploadedOnly")

    def test_settings_api_get_and_put(self):
        # 初期状態: 空の設定
        response = self.client.get("/api/settings/")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["csv_path"], "")
        self.assertEqual(data["inspection_folder_paths"], [])
        self.assertEqual(data["inspection_folder_priorities"], {})

        # PUTで保存
        put_response = self.client.put(
            "/api/settings/",
            {
                "csv_path": r"temp\master.csv",
                "inspection_folder_paths": [
                    r"\\server\share\自動機検査書フォルダ1",
                    r"\\server\share\製品検査(1)フォルダ",
                ],
                "inspection_folder_priorities": {
                    r"\\server\share\自動機検査書フォルダ1": 100,
                    r"\\server\share\製品検査(1)フォルダ": 10,
                    r"\\server\share\削除済み": 999,
                },
            },
            format="json",
        )
        self.assertEqual(put_response.status_code, 200)

        get_response = self.client.get("/api/settings/")
        self.assertEqual(get_response.status_code, 200)
        data2 = get_response.json()
        self.assertEqual(data2["csv_path"], r"temp\master.csv")
        self.assertEqual(len(data2["inspection_folder_paths"]), 2)
        self.assertEqual(
            data2["inspection_folder_priorities"],
            {
                r"\\server\share\自動機検査書フォルダ1": 100,
                r"\\server\share\製品検査(1)フォルダ": 10,
            },
        )

        old_format_response = self.client.put(
            "/api/settings/",
            {"inspection_folder_paths": [r"\\server\share\自動機検査書フォルダ1", r"\\server\share\追加"]},
            format="json",
        )
        self.assertEqual(old_format_response.status_code, 200)
        self.assertEqual(
            old_format_response.json()["inspection_folder_priorities"],
            {r"\\server\share\自動機検査書フォルダ1": 100, r"\\server\share\追加": 0},
        )

    def test_settings_api_rejects_non_integer_priority(self):
        response = self.client.put(
            "/api/settings/",
            {
                "inspection_folder_paths": [r"\\server\share\folder"],
                "inspection_folder_priorities": {r"\\server\share\folder": "high"},
            },
            format="json",
        )
        self.assertEqual(response.status_code, 400)

    def test_settings_api_trims_paths_and_matches_priority_keys_case_insensitively(self):
        response = self.client.put(
            "/api/settings/",
            {
                "inspection_folder_paths": [r"  C:\Inspection\Primary  "],
                "inspection_folder_priorities": {r" c:\inspection\primary ": 100},
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["inspection_folder_paths"], [r"C:\Inspection\Primary"])
        self.assertEqual(
            response.json()["inspection_folder_priorities"],
            {r"C:\Inspection\Primary": 100},
        )

    def test_settings_api_rejects_blank_and_normalized_duplicate_paths(self):
        blank_response = self.client.put(
            "/api/settings/",
            {"inspection_folder_paths": ["   "]},
            format="json",
        )
        self.assertEqual(blank_response.status_code, 400)
        self.assertIn("空のフォルダパス", str(blank_response.json()))

        duplicate_response = self.client.put(
            "/api/settings/",
            {
                "inspection_folder_paths": [
                    r"C:\Inspection\Primary",
                    r" c:/inspection/primary/ ",
                ],
            },
            format="json",
        )
        self.assertEqual(duplicate_response.status_code, 400)
        self.assertIn("同じフォルダパス", str(duplicate_response.json()))

    def test_settings_old_put_preserves_priority_after_trim_and_case_change(self):
        initial_response = self.client.put(
            "/api/settings/",
            {
                "inspection_folder_paths": [r"C:\Inspection\Primary"],
                "inspection_folder_priorities": {r"C:\Inspection\Primary": 75},
            },
            format="json",
        )
        self.assertEqual(initial_response.status_code, 200)

        old_format_response = self.client.put(
            "/api/settings/",
            {"inspection_folder_paths": [r"  c:\inspection\primary  "]},
            format="json",
        )

        self.assertEqual(old_format_response.status_code, 200)
        self.assertEqual(old_format_response.json()["inspection_folder_paths"], [r"c:\inspection\primary"])
        self.assertEqual(
            old_format_response.json()["inspection_folder_priorities"],
            {r"c:\inspection\primary": 75},
        )

    def test_class_4_press_from_node_type_1(self):
        csv_lines = [CSV_HEADER, make_csv_row("BA0080", "CAD0001", "DK0289", "Press部品", 5, "工場", 1)]
        csv_bytes = "\n".join(csv_lines).encode("cp932")
        response = self._upload_csv(csv_bytes)
        self.assertEqual(response.status_code, 202)

        dk0289 = Master.objects.get(code="DK0289")
        self.assertEqual(dk0289.node_type_1, "プレス")

        mc = MasterClass.objects.select_related("class_master").get(master=dk0289)
        self.assertEqual(mc.class_master.class_no, 4)

    def test_class_5_secondary_processing(self):
        master = Master.objects.create(
            code="CAZ0001", name="加工品", node_type_1="加工", department="製造管理部"
        )
        master2 = Master.objects.create(
            code="CAZ0002", name="加工品2", node_type_1="加工", department="生残技術部"
        )
        master3 = Master.objects.create(
            code="CAZ0003", name="加工品他部署", node_type_1="加工", department="営業部"
        )

        from quality.services import determine_inspection_class
        self.assertEqual(determine_inspection_class(master, {}), 5)
        self.assertEqual(determine_inspection_class(master2, {}), 5)
        self.assertEqual(determine_inspection_class(master3, {}), 8)

    def test_class_1_from_machine_assignment(self):
        master = Master.objects.create(code="CAG0999", name="自動機対象")
        machine = Machine.objects.create(
            machine_no="AUTO-TEST-1",
            machine_name="自動機テスト",
            machine_class=1,
            shape_type=Machine.ShapeType.RECTANGLE,
            map_x=0,
            map_y=0,
            width=1,
            height=1,
        )
        MachineAssignment.objects.create(machine=machine, code=master, assignment_class=1)

        from quality.services import determine_inspection_class
        self.assertEqual(determine_inspection_class(master, {}), 1)

    def test_class_8_fallback(self):
        master = Master.objects.create(
            code="ZZZ9999", name="該当なし"
        )
        from quality.services import determine_inspection_class
        self.assertEqual(determine_inspection_class(master, {}), 8)

    def test_import_result_structure(self):
        response = self._upload_csv(MASTER_CSV)
        self.assertEqual(response.status_code, 202)

        job_id = response.json()["job_id"]
        job = Job.objects.get(job_id=job_id)
        self.assertEqual(job.status, Job.Status.SUCCEEDED)

        result = job.result
        self.assertIn("updated_master_count", result)
        self.assertIn("updated_class_count", result)
        self.assertIn("updated_structure_count", result)
        self.assertIn("inspection_file_count", result)
        self.assertIn("source", result)

        self.assertGreater(result["updated_master_count"], 0)
        self.assertGreater(result["updated_structure_count"], 0)
        self.assertGreater(result["updated_class_count"], 0)

    def test_classify_master_by_product_code_function(self):
        from quality.services import classify_master_by_product_code

        # BA prefix -> スライド丁番
        r1 = classify_master_by_product_code("BA0061")
        self.assertEqual(r1["category"], "スライド丁番")
        self.assertEqual(r1["node_type_1"], "")
        self.assertEqual(r1["node_type_2"], "")

        # CAG -> 鍍金
        r2 = classify_master_by_product_code("CAG0008")
        self.assertEqual(r2["node_type_1"], "鍍金")
        self.assertEqual(r2["category"], "スライド丁番")

        # CAL -> 加工, タッピング
        r3 = classify_master_by_product_code("CAL0001")
        self.assertEqual(r3["node_type_1"], "加工")
        self.assertEqual(r3["node_type_2"], "タッピング")

        # DA -> ダイカスト
        r4 = classify_master_by_product_code("DA0045")
        self.assertEqual(r4["node_type_1"], "ダイカスト")

        # 不明なコード
        r5 = classify_master_by_product_code("XX9999")
        self.assertEqual(r5["node_type_1"], "")
        self.assertEqual(r5["node_type_2"], "")
        self.assertEqual(r5["category"], "")

        # 最長一致: CAD should match CAD, not CA or C
        r6 = classify_master_by_product_code("CAD0001")
        self.assertEqual(r6["node_type_1"], "加工")
        self.assertEqual(r6["node_type_2"], "焼入れ")

    def test_inspection_sheet_required_uses_master_class(self):
        master = Master.objects.create(code="MCLASS1", name="Class1")
        MasterClass.objects.create(master=master, class_master=ClassMaster.objects.get(class_no=1))

        master6 = Master.objects.create(code="MCLASS6", name="Class6")
        MasterClass.objects.create(master=master6, class_master=ClassMaster.objects.get(class_no=6))

        master7 = Master.objects.create(code="MCLASS7", name="Class7")
        MasterClass.objects.create(master=master7, class_master=ClassMaster.objects.get(class_no=7))

        master2 = Master.objects.create(code="MCLASS2", name="Class2")
        MasterClass.objects.create(master=master2, class_master=ClassMaster.objects.get(class_no=2))

        from quality.services import inspection_sheet_required
        self.assertTrue(inspection_sheet_required(master))
        self.assertTrue(inspection_sheet_required(master6))
        self.assertTrue(inspection_sheet_required(master7))
        self.assertFalse(inspection_sheet_required(master2))
        self.assertFalse(inspection_sheet_required(None))
        self.assertFalse(inspection_sheet_required(Master.objects.create(code="NOCLASS", name="NoClass")))

    def test_targets_include_product_category(self):
        master = Master.objects.create(
            code="CAP0048", name="テスト品名",
            product_category="スライド丁番",
        )
        machine = Machine.objects.create(machine_no="CAT-M", machine_name="Category", machine_class=1, shape_type="rectangle", map_x=0, map_y=0, width=1, height=1)
        MachineAssignment.objects.create(machine=machine, code=master, assignment_class=1)

        self.client.post(
            "/api/inspection-targets/factory-map/",
            {"date": "2026-06-01", "machine_id": machine.id, "code": "CAP0048"},
            format="json",
        )
        targets = self.client.get("/api/inspection-targets/?date=2026-06-01").json()
        self.assertEqual(len(targets), 1)
        t = targets[0]
        self.assertEqual(t["product_category"], "スライド丁番")
        self.assertEqual(t["category"], 1)  # MasterClass 由来

    def test_settings_persist_after_save(self):
        self.client.put(
            "/api/settings/",
            {"csv_path": "persist_test.csv", "inspection_folder_paths": ["/path/a", "/path/b"]},
            format="json",
        )

        # 別のリクエストとして再取得
        response = self.client.get("/api/settings/")
        data = response.json()
        self.assertEqual(data["csv_path"], "persist_test.csv")
        self.assertEqual(data["inspection_folder_paths"], ["/path/a", "/path/b"])

    def test_master_update_without_file_and_without_settings_falls_back_to_default(self):
        small_master_csv = (CSV_HEADER + "\n" + MASTER_CSV_LINES[0]).encode("cp932")
        with TemporaryDirectory() as temp_dir:
            Path(temp_dir, "master.csv").write_bytes(small_master_csv)
            with override_settings(TEST_INPUT_DIR=Path(temp_dir)):
                response = self.client.post(
                    "/api/master/update/",
                    {"force": False},
                    format="json",
                )
        self.assertEqual(response.status_code, 202)
        self.assertTrue(Master.objects.filter(code="BA0061", name="DYL完成品").exists())

    def test_excel_serial_date(self):
        from quality.services import _excel_serial_date
        from datetime import date
        self.assertEqual(_excel_serial_date(date(1900, 1, 1)), 2)
        self.assertEqual(_excel_serial_date(date(2026, 6, 26)), 46199)

    def test_issue_inspection_sheets_raises_on_missing_template(self):
        from quality.services import issue_inspection_sheets
        from django.test import override_settings
        from pathlib import Path
        with self.settings(DAILY_REPORT_TEMPLATE=Path("C:/nonexistent_template_path.xlsm")):
            with self.assertRaises(FileNotFoundError):
                issue_inspection_sheets()

    def test_issue_inspection_sheets_builds_correct_rows(self):
        from quality.services import issue_inspection_sheets, _excel_serial_date
        from datetime import date
        from unittest.mock import MagicMock, patch, PropertyMock

        cm = ClassMaster.objects.get(class_no=1)
        master = Master.objects.create(code="CDP0028", name="TestItem")
        MasterClass.objects.create(master=master, class_master=cm)
        InspectionFile.objects.create(master=master, file_name="test.xls", file_path=r"C:\test\test.xls")
        h = History.objects.create(date=date(2026, 6, 26), master=master, time_slot="A")

        cells_data = {}

        def make_cell():
            cell = MagicMock()
            cell_mock_value = MagicMock()
            def set_value(val):
                cell_mock_value.Value = val
            cell.Value = cell_mock_value
            type(cell).Value = PropertyMock(side_effect=lambda: cell_mock_value.Value)
            return cell

        cell_map = {}

        def track_cells(row, col):
            key = (row, col)
            if key not in cell_map:
                cell = MagicMock()
                cell_mock = MagicMock()
                def set_val(v):
                    cell_mock.Value = v
                def get_val():
                    if hasattr(cell_mock, '_value'):
                        return cell_mock._value
                    return None
                cell_mock._value = None
                cell_mock.Value = None
                cell_mock.configure_mock(**{'Value': cell_mock.Value})
                cell.Value = None
                cell_map[key] = cell
            return cell_map[key]

        with TemporaryDirectory() as tmp:
            template_path = Path(tmp) / "daily.xlsm"
            workbook = Workbook()
            workbook.save(template_path)

            mock_ws = MagicMock()
            mock_ws.Name = "data"
            mock_ws.Cells.side_effect = track_cells

            mock_wb = MagicMock()
            mock_wb.Sheets.return_value = mock_ws
            mock_wb.Sheets.__iter__.return_value = iter([mock_ws])

            mock_xl = MagicMock()
            mock_xl.Workbooks.Open.return_value = mock_wb

            with mock_excel_com_modules() as (pythoncom, win32_client):
                with patch.object(
                    win32_client,
                    "GetActiveObject",
                    side_effect=pythoncom.com_error(-2147023584, "No active Excel instance", None, None),
                ), patch.object(win32_client, "Dispatch", return_value=mock_xl) as mocked_dispatch:
                    with override_settings(DAILY_REPORT_TEMPLATE=template_path):
                        result = issue_inspection_sheets(target_date=date(2026, 6, 26))

            self.assertEqual(result["issued_count"], 1)
            mocked_dispatch.assert_called_once_with("Excel.Application")
            mock_xl.Workbooks.Open.assert_called_once()

            cells_calls = mock_ws.Cells.call_args_list
            row_col_calls = [ca[0] for ca in cells_calls]
            self.assertIn((2, 1), row_col_calls)
            self.assertIn((2, 3), row_col_calls)
            self.assertIn((2, 4), row_col_calls)
            self.assertIn((2, 5), row_col_calls)
            self.assertIn((2, 6), row_col_calls)

            h.refresh_from_db()
            self.assertTrue(h.is_sheet_issued)

    def test_issue_inspection_sheets_keeps_shared_excel_open(self):
        from quality.services import issue_inspection_sheets
        from datetime import date
        from unittest.mock import MagicMock, patch

        cm = ClassMaster.objects.get(class_no=1)
        master = Master.objects.create(code="CDP0029", name="Shared Excel item")
        MasterClass.objects.create(master=master, class_master=cm)
        InspectionFile.objects.create(master=master, file_name="test.xls", file_path=r"C:\test\test.xls")
        History.objects.create(date=date(2026, 6, 26), master=master, time_slot="A")

        with TemporaryDirectory() as tmp:
            template_path = Path(tmp) / "daily.xlsm"
            Workbook().save(template_path)

            mock_ws = MagicMock()
            mock_ws.Name = "data"
            mock_wb = MagicMock()
            mock_wb.Sheets.return_value = mock_ws
            mock_wb.Sheets.__iter__.return_value = iter([mock_ws])
            shared_xl = MagicMock()
            shared_xl.Workbooks.Open.return_value = mock_wb

            with mock_excel_com_modules() as (_, win32_client):
                with patch.object(win32_client, "GetActiveObject", return_value=shared_xl) as active_excel, \
                     patch.object(win32_client, "Dispatch") as mocked_dispatch, \
                     override_settings(DAILY_REPORT_TEMPLATE=template_path):
                    result = issue_inspection_sheets(target_date=date(2026, 6, 26))

            self.assertEqual(result["issued_count"], 1)
            active_excel.assert_called_once_with("Excel.Application")
            mocked_dispatch.assert_not_called()
            mock_wb.Close.assert_called_once_with(False)
            shared_xl.Quit.assert_not_called()

    def test_issue_inspection_sheets_filters_non_target_class(self):
        from quality.services import issue_inspection_sheets
        from datetime import date
        from unittest.mock import MagicMock, patch

        cm = ClassMaster.objects.get(class_no=4)
        master = Master.objects.create(code="TEST01", name="NonTarget")
        MasterClass.objects.create(master=master, class_master=cm)
        History.objects.create(date=date(2026, 6, 26), master=master, time_slot="A")

        with TemporaryDirectory() as tmp:
            template_path = Path(tmp) / "daily.xlsm"
            workbook = Workbook()
            workbook.save(template_path)

            mock_ws = MagicMock()
            mock_ws.Name = "data"

            mock_wb = MagicMock()
            mock_wb.Sheets.return_value = mock_ws
            mock_wb.Sheets.__iter__.return_value = iter([mock_ws])

            mock_xl = MagicMock()
            mock_xl.Workbooks.Open.return_value = mock_wb

            with mock_excel_com_modules() as (_, win32_client):
                with patch.object(win32_client, "Dispatch", return_value=mock_xl):
                    with override_settings(DAILY_REPORT_TEMPLATE=template_path):
                        result = issue_inspection_sheets(target_date=date(2026, 6, 26))

            self.assertEqual(result["issued_count"], 0)
            self.assertIn("No printable entries", result["message"])

    def test_issue_inspection_sheets_skips_missing_file_path(self):
        from quality.services import issue_inspection_sheets
        from datetime import date
        from unittest.mock import MagicMock, patch

        cm = ClassMaster.objects.get(class_no=1)
        master = Master.objects.create(code="TEST02", name="NoFileItem")
        MasterClass.objects.create(master=master, class_master=cm)
        History.objects.create(date=date(2026, 6, 26), master=master, time_slot="A")

        with TemporaryDirectory() as tmp:
            template_path = Path(tmp) / "daily.xlsm"
            workbook = Workbook()
            workbook.save(template_path)

            mock_ws = MagicMock()
            mock_ws.Name = "data"

            mock_wb = MagicMock()
            mock_wb.Sheets.return_value = mock_ws
            mock_wb.Sheets.__iter__.return_value = iter([mock_ws])

            mock_xl = MagicMock()
            mock_xl.Workbooks.Open.return_value = mock_wb

            with mock_excel_com_modules() as (_, win32_client):
                with patch.object(win32_client, "Dispatch", return_value=mock_xl):
                    with override_settings(DAILY_REPORT_TEMPLATE=template_path):
                        result = issue_inspection_sheets(target_date=date(2026, 6, 26))

            self.assertEqual(result["issued_count"], 0)
            self.assertIn("No printable entries", result["message"])

    def test_utf8_sig_csv_import(self):
        """utf-8-sig BOM付きCSVでも読み込めること"""
        csv_bytes = ("\ufeff" + CSV_HEADER + "\n" + "\n".join(MASTER_CSV_LINES)).encode("utf-8-sig")
        response = self._upload_csv(csv_bytes)
        self.assertEqual(response.status_code, 202)
        self.assertTrue(Master.objects.filter(code="BA0061").exists())


class InspectionFileViewTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = authenticate_test_client(self.client)
        self.master = Master.objects.create(code="PDF001", name="PDF test")

    def _create_inspection_file(self, path):
        return InspectionFile.objects.create(master=self.master, file_name=Path(path).name, file_path=str(path))

    def test_pdf_endpoint_streams_converted_file_and_cleans_it_up(self):
        with TemporaryDirectory() as tmp:
            source_path = Path(tmp) / "PDF001.xlsx"
            converted_path = Path(tmp) / "converted.pdf"
            source_path.write_bytes(b"spreadsheet")
            converted_path.write_bytes(b"%PDF-test")
            self._create_inspection_file(source_path)
            with patch("quality.services.convert_excel_to_pdf", return_value=str(converted_path)):
                response = self.client.get("/api/inspection-file/pdf/?code=PDF001")
            self.assertEqual(response.status_code, 200)
            self.assertEqual(response["Content-Type"], "application/pdf")
            self.assertEqual(b"".join(response.streaming_content), b"%PDF-test")
            response.close()
            self.assertFalse(converted_path.exists())

    def test_pdf_endpoint_returns_conversion_error(self):
        with TemporaryDirectory() as tmp:
            source_path = Path(tmp) / "PDF001.xlsx"
            source_path.write_bytes(b"spreadsheet")
            self._create_inspection_file(source_path)
            with patch("quality.services.convert_excel_to_pdf", side_effect=RuntimeError("conversion failed")):
                response = self.client.get("/api/inspection-file/pdf/?code=PDF001")
            self.assertEqual(response.status_code, 500)
            self.assertEqual(response.json()["error_code"], "CONVERT_FAILED")

    def test_open_endpoint_returns_open_failed(self):
        with TemporaryDirectory() as tmp:
            source_path = Path(tmp) / "PDF001.xlsx"
            source_path.write_bytes(b"spreadsheet")
            self._create_inspection_file(source_path)
            with patch("quality.views.os.startfile", side_effect=OSError("no association"), create=True):
                response = self.client.get("/api/inspection-file/open/?code=PDF001")
            self.assertEqual(response.status_code, 500)
            self.assertEqual(response.json()["error_code"], "OPEN_FAILED")


class JobServiceTests(TestCase):
    def test_run_job_persists_json_safe_failure_result(self):
        from quality.services import run_job

        job = Job.objects.create(
            job_id="job_test_failure",
            job_type=Job.JobType.INSPECTION_SHEET_ISSUE,
            request_payload={},
        )

        with self.assertRaisesRegex(RuntimeError, "Excel unavailable"):
            run_job(job, lambda: (_ for _ in ()).throw(RuntimeError("Excel unavailable")))

        job.refresh_from_db()
        self.assertEqual(job.status, Job.Status.FAILED)
        self.assertEqual(job.error_message, "Excel unavailable")
        self.assertEqual(
            job.result,
            {
                "status": "failed",
                "error_message": "Excel unavailable",
                "exception_type": "RuntimeError",
            },
        )

    def test_run_job_persists_safe_classification_details(self):
        from quality.services import ClassificationError, run_job

        job = Job.objects.create(
            job_id="job_test_classification_failure",
            job_type=Job.JobType.INSPECTION_SHEET_ISSUE,
            request_payload={},
        )
        error = ClassificationError(
            "AMBIGUOUS_INSPECTION_FILE",
            "同じクラスの検査書が複数あります。保存場所を確認してください。",
            {
                "code": "CCP0030",
                "class": 1,
                "candidate_count": 2,
                "candidate_file_names": [r"C:\internal\CCP0030-A.xlsx", r"D:\secret\CCP0030-B.xlsx"],
                "internal_path": r"C:\internal",
            },
        )

        with self.assertRaises(ClassificationError):
            run_job(job, lambda: (_ for _ in ()).throw(error))

        job.refresh_from_db()
        self.assertEqual(job.status, Job.Status.FAILED)
        self.assertEqual(job.result["error_code"], "AMBIGUOUS_INSPECTION_FILE")
        self.assertEqual(
            job.result["details"],
            {
                "code": "CCP0030",
                "class": 1,
                "candidate_count": 2,
                "candidate_file_names": ["CCP0030-A.xlsx", "CCP0030-B.xlsx"],
            },
        )
        self.assertIn("品番: CCP0030", job.error_message)
        self.assertIn("候補: 2件", job.error_message)
        self.assertNotIn("internal_path", str(job.result))
        self.assertNotIn("C:\\internal", str(job.result))
        self.assertNotIn("D:\\secret", job.error_message)


class InspectionSummaryApiTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = authenticate_test_client(self.client, "summary-admin")
        self.worker = User.objects.create(login_name="summary-worker", display_name="検査者", password_hash="!", role=User.Role.WORKER)
        self.master = Master.objects.create(code="SUM001", name="集計,品目")
        class_master, _ = ClassMaster.objects.get_or_create(class_no=1, defaults={"class_name": "自動機"})
        MasterClass.objects.create(master=self.master, class_master=class_master)

    def test_note_is_separated_by_owner_and_created_without_targets(self):
        self.client.force_authenticate(self.worker)
        response = self.client.put("/api/inspection-note/", {"date": "2026-07-01", "note": "気づき"}, format="json")
        self.assertEqual(response.status_code, 200)
        session = InspectionSession.objects.get(owner_user=self.worker, target_date="2026-07-01")
        self.assertEqual(session.note, "気づき")
        session.status = InspectionSession.Status.CLOSED
        session.deleted_at = "2026-07-01T01:00:00Z"
        session.deleted_by = self.admin
        session.save(update_fields=["status", "deleted_at", "deleted_by"])
        self.client.put("/api/inspection-note/", {"date": "2026-07-01", "note": "追記"}, format="json")
        session.refresh_from_db()
        self.assertEqual(session.note, "追記")
        self.assertIsNone(session.deleted_at)
        self.assertIsNone(session.deleted_by)
        self.assertEqual(self.client.get("/api/inspection-note/?date=2026-07-01").json()["note"], "追記")
        other = User.objects.create(login_name="other-worker", display_name="別担当", password_hash="!", role=User.Role.WORKER)
        self.client.force_authenticate(other)
        self.assertEqual(self.client.get("/api/inspection-note/?date=2026-07-01").json()["note"], "")

    def test_summary_requires_admin_and_counts_override_excluding_deleted(self):
        History.objects.create(date="2026-07-01", master=self.master, time_slot="A", created_by=self.worker)
        History.objects.create(date="2026-07-01", master=self.master, time_slot="B", class_override=8, created_by=self.worker)
        History.objects.create(date="2026-07-01", master=self.master, time_slot="C", created_by=self.admin, deleted_at="2026-07-01T01:00:00Z", deleted_by=self.admin)
        self.client.force_authenticate(self.worker)
        self.assertEqual(self.client.get("/api/inspection-summary/?start=2026-07-01&end=2026-07-02").status_code, 403)
        self.client.force_authenticate(self.admin)
        response = self.client.get("/api/inspection-summary/?start=2026-07-01&end=2026-07-02")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["class_totals"]["1"], 1)
        self.assertEqual(data["class_totals"]["8"], 1)
        self.assertEqual([day["total"] for day in data["days"]], [2, 0])
        self.assertEqual(data["days"][0]["inspectors"][0]["name"], "検査者")

    def test_summary_keeps_users_with_same_display_name_separate(self):
        other = User.objects.create(login_name="same-name-worker", display_name="検査者", password_hash="!", role=User.Role.WORKER)
        History.objects.create(date="2026-07-01", master=self.master, time_slot="A", created_by=self.worker)
        History.objects.create(date="2026-07-01", master=self.master, time_slot="B", created_by=other)
        InspectionSession.objects.create(target_date="2026-07-01", owner_user=self.worker, note="一人目")
        InspectionSession.objects.create(target_date="2026-07-01", owner_user=other, note="二人目")
        response = self.client.get("/api/inspection-summary/?start=2026-07-01&end=2026-07-01")
        inspectors = response.json()["days"][0]["inspectors"]
        self.assertEqual(len(inspectors), 2)
        self.assertEqual({row["user_id"] for row in inspectors}, {self.worker.user_id, other.user_id})
        self.assertEqual({row["total"] for row in inspectors}, {1})
        notes = response.json()["days"][0]["notes"]
        self.assertEqual({row["user_id"] for row in notes}, {self.worker.user_id, other.user_id})

    def test_top_items_filters_inspectors_and_classes_together(self):
        other = User.objects.create(login_name="same-name-worker", display_name=self.worker.display_name, password_hash="!", role=User.Role.WORKER)
        second_master = Master.objects.create(code="SUM002", name="別品目")
        second_class, _ = ClassMaster.objects.get_or_create(class_no=2, defaults={"class_name": "半自動機"})
        MasterClass.objects.create(master=second_master, class_master=second_class)
        History.objects.create(date="2026-07-01", master=self.master, time_slot="A", created_by=self.worker)
        History.objects.create(date="2026-07-01", master=self.master, time_slot="B", created_by=other)
        History.objects.create(date="2026-07-01", master=second_master, time_slot="C", created_by=self.worker)
        History.objects.create(date="2026-07-01", master=second_master, time_slot="D", created_by=None)

        base = "/api/inspection-summary/?start=2026-07-01&end=2026-07-01"
        unfiltered = self.client.get(base).json()["top_items"]
        self.assertEqual({row["code"]: row["count"] for row in unfiltered}, {"SUM001": 2, "SUM002": 2})

        filtered = self.client.get(f"{base}&classes=1&inspectors={self.worker.user_id}").json()["top_items"]
        self.assertEqual(filtered, [{"code": "SUM001", "name": self.master.name, "count": 1}])
        self.assertEqual(self.client.get(f"{base}&inspectors=").json()["top_items"], [])

        same_name = self.client.get(f"{base}&inspectors={other.user_id}").json()["top_items"]
        self.assertEqual(same_name, [{"code": "SUM001", "name": self.master.name, "count": 1}])
        unknown = self.client.get(f"{base}&inspectors=unknown").json()["top_items"]
        self.assertEqual(unknown, [{"code": "SUM002", "name": second_master.name, "count": 1}])

    def test_top_items_rejects_invalid_inspector_token(self):
        response = self.client.get("/api/inspection-summary/?start=2026-07-01&end=2026-07-01&inspectors=invalid")
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()["error_code"], "INVALID_INSPECTORS")

    def test_csv_has_bom_quotes_and_invalid_period_is_rejected(self):
        InspectionSession.objects.create(target_date="2026-07-01", owner_user=self.worker, note='改行\n"引用",カンマ')
        response = self.client.get("/api/inspection-summary/csv/notes/?start=2026-07-01&end=2026-07-01")
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.content.startswith(b"\xef\xbb\xbf"))
        self.assertIn(b'""', response.content)
        count_csv = self.client.get("/api/inspection-summary/csv/counts/?start=2026-07-01&end=2026-07-02")
        self.assertEqual(count_csv.status_code, 200)
        self.assertIn("2026-07-02,0,0,0,0,0,0,0,0,0,0", count_csv.content.decode("utf-8-sig"))
        self.assertEqual(self.client.get("/api/inspection-summary/?start=2026-07-02&end=2026-07-01").status_code, 400)
