import csv
import hashlib
import os
import shutil
import logging
import re
import subprocess
import sys
from time import perf_counter
from pathlib import Path
from uuid import uuid4

from django.conf import settings
from django.db import connection, transaction
from django.db.models import Q
from django.utils import timezone
import fitz
from openpyxl import load_workbook
from pypdf import PdfReader
from rapidocr_onnxruntime import RapidOCR
import xlrd

from datetime import date as date_type, datetime
from .constants import CHECK_SLOTS
from .s2_cr08_shared import (
    ADVISORY_LOCK_NAMESPACE,
    make_advisory_lock_id,
    make_application_name_marker,
)
from .models import (
    ClassMaster,
    History,
    InspectionFile,
    InspectionSession,
    InspectionTarget,
    InspectionTargetWarning,
    Job,
    MachineAssignment,
    Master,
    MasterClass,
    SpecialInspectionClass9,
    Structure,
)


CODE_PATTERN = re.compile(r"\b[A-Za-z]{3}\d{4}\b")
FILE_CODE_PATTERN = re.compile(r"(?<![A-Za-z0-9])[A-Za-z]{2,4}\d{4}(?![A-Za-z0-9])")
ALNUM_PATTERN = re.compile(r"[A-Za-z0-9]")
INSPECTION_FOLDER_SCAN_ERROR = "Inspection folder scan failed or became inaccessible."
logger = logging.getLogger(__name__)
_OCR_ENGINE = None


def normalize_code(code):
    return str(code).strip().upper()


def inspection_sheet_required(master):
    if master is None:
        return False
    mc = MasterClass.objects.filter(master=master).exclude(class_master__class_no=9).first()
    return mc is not None and mc.class_master is not None and mc.class_master.class_no in (1, 6, 7)


class ClassificationError(Exception):
    def __init__(self, error_code, message, details=None):
        super().__init__(message)
        self.error_code = error_code
        self.message = message
        self.details = details or {}


def inspection_sheet_required_for_class(class_no):
    return class_no in (1, 6, 7, 9)


# PRODUCT.md classification table
# (prefix, node_type_1, node_type_2, category)
PRODUCT_CLASSIFICATION = [
    ("BA", "", "", "スライド丁番"),
    ("BB", "", "", "丁番"),
    ("BC", "", "", "戸車"),
    ("BD", "", "", "吊り車"),
    ("BE", "", "", "キャッチ"),
    ("BF", "", "", "取手・つまみ・引き手"),
    ("BG", "", "", "手摺り"),
    ("BH", "", "", "レール"),
    ("BI", "", "", "カーテン・ブラインド金物"),
    ("BJ", "", "", ""),
    ("BK", "", "", "意匠系金物"),
    ("BL", "", "", "機能金物"),
    ("BM", "", "", "固定金物"),
    ("BN", "", "", "諸金物（その他）"),
    ("BO", "", "", "付属金物"),
    ("BP", "", "", "部品類"),
    ("BQ", "", "", ""),
    ("BR", "", "", ""),
    ("BS", "", "", ""),
    ("BT", "", "", "ラック"),
    ("BU", "", "", "POP"),
    ("BV", "", "", ""),
    ("BW", "", "", ""),
    ("BX", "", "", ""),
    ("BY", "", "", "提案品"),
    ("BZ", "", "", "その他製品"),
    ("CAA", "プレス", "", "スライド丁番"),
    ("CAD", "加工", "焼入れ", "スライド丁番"),
    ("CAE", "加工", "バレル", "スライド丁番"),
    ("CAG", "鍍金", "", "スライド丁番"),
    ("CAH", "塗装", "", "スライド丁番"),
    ("CAI", "組立", "", "スライド丁番"),
    ("CAK", "加工", "カシメ", "スライド丁番"),
    ("CAL", "加工", "タッピング", "スライド丁番"),
    ("CAM", "組立", "バネ入れ", "スライド丁番"),
    ("CAP", "組立", "自動機", "スライド丁番"),
    ("CAR", "袋入", "", "スライド丁番"),
    ("CAS", "梱包", "", "スライド丁番"),
    ("CAZ", "加工", "その他", "スライド丁番"),
    ("CBA", "プレス", "", "丁番"),
    ("CBB", "加工", "切断", "丁番"),
    ("CBC", "加工", "溶接", "丁番"),
    ("CBD", "加工", "焼入れ", "丁番"),
    ("CBE", "加工", "バレル", "丁番"),
    ("CBF", "加工", "皿もみ", "丁番"),
    ("CBG", "鍍金", "", "丁番"),
    ("CBH", "塗装", "", "丁番"),
    ("CBI", "組立", "", "丁番"),
    ("CBJ", "加工", "圧入", "丁番"),
    ("CBK", "加工", "カシメ", "丁番"),
    ("CBL", "加工", "タップ", "丁番"),
    ("CBO", "プレス", "ダボ出し", "丁番"),
    ("CBP", "組立", "自動機", "丁番"),
    ("CBQ", "検査", "", "丁番"),
    ("CBR", "袋入", "", "丁番"),
    ("CBS", "梱包", "", "丁番"),
    ("CBZ", "加工", "その他", "丁番"),
    ("CCA", "プレス", "", "戸車"),
    ("CCG", "鍍金", "", "戸車"),
    ("CCH", "塗装", "", "戸車"),
    ("CCI", "組立", "", "戸車"),
    ("CCK", "加工", "カシメ", "戸車"),
    ("CCL", "加工", "タップ", "戸車"),
    ("CCP", "組立", "自動機", "戸車"),
    ("CCS", "梱包", "", "戸車"),
    ("CCZ", "加工", "その他", "戸車"),
    ("CDA", "プレス", "", "吊車"),
    ("CDB", "加工", "切削", "吊車"),
    ("CDC", "加工", "溶接", "和金具"),
    ("CDD", "加工", "焼入れ", "吊車"),
    ("CDE", "加工", "研磨", "和金具"),
    ("CDF", "加工", "皿もみ", "吊車"),
    ("CDG", "鍍金", "", "吊車"),
    ("CDH", "塗装", "", "吊車"),
    ("CDI", "組立", "", "吊車"),
    ("CDJ", "加工", "圧入", "吊車"),
    ("CDK", "加工", "カシメ", "吊車"),
    ("CDL", "加工", "タップ", "吊車"),
    ("CDN", "加工", "油付け", "吊車"),
    ("CDP", "組立", "自動機", "吊車"),
    ("CDQ", "検査", "", "吊車"),
    ("CDR", "袋入", "", "吊車"),
    ("DCS", "梱包", "", "吊車"),
    ("CDZ", "加工", "その他", "吊車"),
    ("CEA", "プレス", "", "キャッチ"),
    ("CEG", "鍍金", "", "キャッチ"),
    ("CEI", "組立", "", "キャッチ"),
    ("CEL", "加工", "タップ", "キャッチ"),
    ("CER", "袋入", "", "キャッチ"),
    ("CES", "梱包", "", "キャッチ"),
    ("CEZ", "加工", "その他", "キャッチ"),
    ("CFA", "プレス", "", "取手/つまみ/引手"),
    ("CFB", "加工", "切断", "取手/つまみ/引手"),
    ("CFE", "加工", "研磨", "取手/つまみ/引手"),
    ("CFG", "鍍金", "", "取手/つまみ/引手"),
    ("CFH", "塗装", "", "取手/つまみ/引手"),
    ("CFI", "組立", "", "取手/つまみ/引手"),
    ("CFK", "加工", "カシメ", "取手/つまみ/引手"),
    ("CFL", "加工", "タップ", "取手/つまみ/引手"),
    ("CFR", "袋入", "", "取手/つまみ/引手"),
    ("CFS", "梱包", "", "取手/つまみ/引手"),
    ("CFZ", "加工", "その他", "取手/つまみ/引手"),
    ("CGA", "プレス", "", "手摺"),
    ("CGE", "加工", "研磨", "手摺"),
    ("CGF", "加工", "皿もみ", "手摺"),
    ("CGG", "鍍金", "", "手摺"),
    ("CGH", "塗装", "", "手摺"),
    ("CGI", "組立", "", "手摺"),
    ("CGK", "加工", "カシメ", "手摺"),
    ("CGL", "加工", "タップ", "手摺"),
    ("CGQ", "検査", "", "手摺"),
    ("CGR", "袋入", "", "手摺"),
    ("CGS", "梱包", "", "手摺"),
    ("CGZ", "加工", "その他", "手摺"),
    ("CHS", "梱包", "", "レール"),
    ("CIA", "プレス", "", "カーテン/ブラインド金物"),
    ("CIG", "鍍金", "", "カーテン/ブラインド金物"),
    ("CII", "組立", "", "カーテン/ブラインド金物"),
    ("CIS", "梱包", "", "カーテン/ブラインド金物"),
    ("CKS", "梱包", "", "意匠系金物"),
    ("CLA", "プレス", "", "機能金物"),
    ("CLB", "加工", "切断", "機能金物"),
    ("CLC", "加工", "溶接", "機能金物"),
    ("CLE", "加工", "研磨", "機能金物"),
    ("CLF", "加工", "皿もみ", "機能金物"),
    ("CLG", "鍍金", "", "機能金物"),
    ("CLH", "塗装", "", "機能金物"),
    ("CLI", "組立", "", "機能金物"),
    ("CLK", "加工", "カシメ", "機能金物"),
    ("CLL", "加工", "タップ", "機能金物"),
    ("CLS", "梱包", "", "機能金物"),
    ("CLZ", "加工", "その他", "機能金物"),
    ("CMA", "プレス", "", "固定金物"),
    ("CMC", "加工", "溶接", "固定金物"),
    ("CME", "加工", "研磨", "固定金物"),
    ("CMG", "鍍金", "", "固定金物"),
    ("CMH", "塗装", "", "固定金物"),
    ("CMI", "組立", "", "固定金物"),
    ("CMK", "加工", "カシメ", "固定金物"),
    ("CML", "加工", "タッピング", "固定金物"),
    ("CMQ", "検査", "", "固定金物"),
    ("CMR", "袋入", "", "固定金物"),
    ("CMS", "梱包", "", "固定金物"),
    ("CNA", "プレス", "", "諸金物"),
    ("CNB", "加工", "切断", "諸金物"),
    ("CNC", "加工", "溶接", "諸金物"),
    ("CNE", "加工", "研磨", "諸金物"),
    ("CNF", "加工", "皿もみ", "諸金物"),
    ("CNG", "鍍金", "", "諸金物"),
    ("CNH", "塗装", "", "諸金物"),
    ("CNI", "組立", "", "諸金物"),
    ("CNK", "加工", "カシメ", "諸金物"),
    ("CNL", "加工", "タップ", "諸金物"),
    ("CNS", "梱包", "", "諸金物"),
    ("CNZ", "加工", "その他", "諸金物"),
    ("COA", "プレス", "", "付属金物"),
    ("COB", "加工", "切断", "付属金物"),
    ("COC", "加工", "溶接", "付属金物"),
    ("COD", "加工", "焼入れ", "付属金物"),
    ("COE", "加工", "研磨", "付属金物"),
    ("COF", "加工", "皿もみ", "付属金物"),
    ("COG", "鍍金", "", "付属金物"),
    ("COH", "塗装", "", "付属金物"),
    ("COI", "組立", "", "付属金物"),
    ("COK", "加工", "カシメ", "付属金物"),
    ("COL", "加工", "タップ", "付属金物"),
    ("COR", "袋入", "", "付属金物"),
    ("COS", "梱包", "", "付属金物"),
    ("COZ", "加工", "その他", "付属金物"),
    ("CPA", "プレス", "", "部品類"),
    ("CPB", "加工", "切断", "部品類"),
    ("CPD", "加工", "焼入れ", "部品類"),
    ("CPG", "鍍金", "", "部品類"),
    ("CPH", "塗装", "", "部品類"),
    ("CPR", "袋入", "", "部品類"),
    ("CPS", "梱包", "", "部品類"),
    ("CPZ", "加工", "その他", "部品類"),
    ("CT", "", "", "ラック"),
    ("CTH", "塗装", "", "ラック"),
    ("CTS", "梱包", "", "ラック"),
    ("CTZ", "加工", "その他", "ラック"),
    ("CU", "", "", "POP"),
    ("CUS", "梱包", "", "POP"),
    ("CYA", "プレス", "", "提案品"),
    ("CY", "", "", "提案品"),
    ("CYE", "加工", "研磨", "提案品"),
    ("CYG", "鍍金", "", "提案品"),
    ("CYI", "組立", "", "提案品"),
    ("CYS", "梱包", "", "提案品"),
    ("CZ", "", "", "その他製品"),
    ("CZA", "プレス", "", "その他製品"),
    ("CZB", "加工", "切断", "その他製品"),
    ("CZC", "加工", "溶接", "その他製品"),
    ("CZE", "加工", "研磨", "その他製品"),
    ("CZG", "鍍金", "", "その他製品"),
    ("CZH", "塗装", "", "その他製品"),
    ("CZI", "組立", "", "その他製品"),
    ("CZK", "加工", "カシメ", "その他製品"),
    ("CZL", "加工", "タップ", "その他製品"),
    ("CZR", "袋入", "", "その他製品"),
    ("CZS", "梱包", "", "その他製品"),
    ("CZZ", "加工", "その他", "その他製品"),
    ("DA", "ダイカスト", "", ""),
    ("DAF", "塗装", "", ""),
    ("DAG", "鍍金", "", ""),
    ("DAH", "塗装", "", ""),
    ("DAZ", "加工", "その他", ""),
    ("DB", "樹脂", "", ""),
    ("DBB", "加工", "切断", ""),
    ("DBL", "加工", "タップ", ""),
    ("DBS", "梱包", "", ""),
    ("DBT", "樹脂", "", ""),
    ("DBZ", "加工", "その他", ""),
    ("DC", "ネジ", "", ""),
    ("DCZ", "ネジ", "", ""),
    ("DD", "バネ", "", ""),
    ("DE", "ワッシャー", "", ""),
    ("DF", "ピン", "", ""),
    ("DFG", "鍍金", "", ""),
    ("DG", "ベアリング", "", ""),
    ("DGZ", "加工", "その他", ""),
    ("DH", "鋼球", "", ""),
    ("DI", "磁石", "", ""),
    ("DJ", "ヘッダー", "", ""),
    ("DJG", "ヘッダー", "", ""),
    ("DJZ", "加工", "その他", ""),
    ("DK", "プレス", "", ""),
    ("DKG", "鍍金", "", ""),
    ("DKH", "塗装", "", ""),
    ("DKZ", "加工", "その他", ""),
    ("DL", "丁番用軸", "", ""),
    ("DN", "鋳造", "", ""),
    ("DZ", "その他", "", ""),
    ("DZF", "加工", "皿もみ", ""),
    ("DZG", "鍍金", "", ""),
    ("EA", "材料", "鉄", ""),
    ("EB", "材料", "ステンレス", ""),
    ("EC", "材料", "アルミ", ""),
    ("ED", "材料", "真鍮", ""),
    ("EZ", "材料", "青銅", ""),
    ("FA", "ダイカスト", "", ""),
    ("FAA", "加工", "切断", ""),
    ("FAE", "加工", "研磨", ""),
    ("FAF", "加工", "その他", ""),
    ("FAG", "鍍金", "", ""),
    ("FAL", "加工", "その他", ""),
    ("FAZ", "加工", "その他", ""),
    ("FB", "樹脂", "", ""),
    ("FC", "ネジ", "", ""),
    ("FF", "ピン", "", ""),
    ("FFG", "鍍金", "", ""),
    ("FFK", "加工", "カシメ", ""),
    ("FG", "ベアリング", "", ""),
    ("FJ", "ヘッダー", "", ""),
    ("FJZ", "加工", "その他", ""),
    ("FK", "プレス", "", ""),
    ("FKA", "プレス", "", ""),
    ("FKK", "加工", "皿もみ", ""),
    ("FKG", "鍍金", "", ""),
    ("FKI", "組立", "", ""),
    ("FZ", "プレス", "", ""),
    ("IZS", "梱包", "", ""),
    ("IZZ", "その他", "ピン", ""),
    ("JA", "その他", "", ""),
    ("JB", "樹脂", "", ""),
    ("JC", "ネジ", "", ""),
    ("JE", "FA-160-KY部品", "", ""),
    ("JF", "FA-160-KY部品", "", ""),
    ("JI", "マグネット", "", ""),
    ("JK", "鍍金", "", ""),
    ("JZ", "その他", "", ""),
    ("KZ", "その他", "", ""),
    ("NA", "支給品", "", ""),
    ("NF", "支給品", "", ""),
    ("NZ", "支給品", "", ""),
    ("PA", "梱包用ビニール袋", "", ""),
    ("PB", "梱包用小箱", "", ""),
    ("PC", "梱包用段ボール", "", ""),
    ("PZ", "その他", "", ""),
    ("SB", "その他", "", ""),
    ("SC", "その他", "", ""),
    ("SD", "その他", "", ""),
    ("SL", "その他", "", ""),
    ("YC", "ビニール袋", "", ""),
    ("YD", "小箱", "", ""),
    ("YE", "段ボール", "", ""),
    ("YZ", "その他", "", ""),
]

# Sort by prefix length descending for longest-match
PRODUCT_CLASSIFICATION_SORTED = sorted(PRODUCT_CLASSIFICATION, key=lambda x: len(x[0]), reverse=True)


def classify_master_by_product_code(code):
    for prefix, n1, n2, cat in PRODUCT_CLASSIFICATION_SORTED:
        if code.startswith(prefix):
            return {"node_type_1": n1, "node_type_2": n2, "category": cat}
    return {"node_type_1": "", "node_type_2": "", "category": ""}


def get_assignment_class_for_master(master):
    if master is None:
        return None
    ma = MachineAssignment.objects.filter(
        code=master, assignment_class__isnull=False
    ).order_by("assignment_class").first()
    return ma.assignment_class if ma else None


def _assignment_classes(master):
    if master is None:
        return set()
    classes = set()
    for assignment_class, machine_class in MachineAssignment.objects.filter(code=master).values_list(
        "assignment_class", "machine__machine_class"
    ):
        value = assignment_class if assignment_class is not None else machine_class
        if value is not None:
            classes.add(int(value))
    return classes


def _conflicting_machine_numbers(master):
    machine_numbers = []
    assignments = MachineAssignment.objects.filter(code=master).select_related("machine").order_by("machine__machine_no")
    for assignment in assignments:
        assignment_class = assignment.assignment_class
        effective_class = assignment_class if assignment_class is not None else assignment.machine.machine_class
        if effective_class in (1, 2):
            machine_numbers.append(assignment.machine.machine_no)
    return machine_numbers


def _safe_file_name(file_path):
    return Path(str(file_path).replace("\\", "/")).name


def resolve_process_class(master):
    if master is None:
        raise ClassificationError("PROCESS_CLASS_NOT_FOUND", "工程分類に必要な品目マスターがありません。")
    classes = _assignment_classes(master)
    process_12 = classes & {1, 2}
    if process_12 == {1, 2}:
        raise ClassificationError(
            "CLASS_1_2_CONFLICT",
            "クラス1と2の機械設定が重複しています。機械設定を確認してください。",
            {
                "code": master.code,
                "detected_classes": [1, 2],
                "machine_numbers": _conflicting_machine_numbers(master),
            },
        )
    if process_12:
        return next(iter(process_12))
    if 3 in classes:
        return 3
    n1 = (master.node_type_1 or "").strip()
    dept = (master.department or "").strip()
    if n1 == "プレス":
        return 4
    if n1 == "加工" and dept in ("製造管理部", "生産技術部"):
        return 5
    raise ClassificationError(
        "PROCESS_CLASS_NOT_FOUND",
        "工程クラスを判定できません。機械または品目設定を確認してください。",
        {"code": master.code},
    )


def resolve_ocr_class(master):
    process_not_found = None
    try:
        return resolve_process_class(master)
    except ClassificationError as exc:
        if exc.error_code != "PROCESS_CLASS_NOT_FOUND":
            raise
        process_not_found = exc
    if MasterClass.objects.filter(master=master, class_master__class_no=8).exists():
        return 8
    raise process_not_found


def resolve_product_inspection_class(master):
    if master is None:
        raise ClassificationError("PRODUCT_INSPECTION_FILE_NOT_FOUND", "製品検査の品目マスターがありません。")
    folder_classes = {
        classify_folder(path)
        for path in InspectionFile.objects.filter(master=master).values_list("file_path", flat=True)
    }
    matched = folder_classes & {"product1", "product2"}
    if matched == {"product1", "product2"}:
        candidate_file_names = [
            _safe_file_name(path)
            for path in InspectionFile.objects.filter(master=master).order_by("id").values_list("file_path", flat=True)
            if classify_folder(path) in matched
        ]
        raise ClassificationError(
            "CLASS_6_7_CONFLICT",
            "製品検査(1)と製品検査(2)の両方に検査書があります。保存場所を確認してください。",
            {
                "code": master.code,
                "detected_classes": [6, 7],
                "candidate_file_names": candidate_file_names,
            },
        )
    if "product1" in matched:
        return 6
    if "product2" in matched:
        return 7
    raise ClassificationError(
        "PRODUCT_INSPECTION_FILE_NOT_FOUND",
        "製品検査用の検査書が見つかりません。",
        {"code": master.code},
    )


def resolve_special_class(master):
    if master is None or not SpecialInspectionClass9.objects.filter(master=master).exists():
        raise ClassificationError(
            "CLASS_9_SETTING_NOT_FOUND",
            "特殊検査の設定が見つかりません。",
            {"code": master.code if master else None},
        )
    return 9


def resolve_class_for_route(master, registration_route):
    if registration_route == InspectionTarget.RegistrationRoute.OCR:
        return resolve_ocr_class(master)
    if registration_route == InspectionTarget.RegistrationRoute.FACTORY_MAP:
        return resolve_process_class(master)
    if registration_route in (InspectionTarget.RegistrationRoute.EXCEL, InspectionTarget.RegistrationRoute.MANUAL_CODE):
        return resolve_product_inspection_class(master)
    if registration_route == InspectionTarget.RegistrationRoute.SPECIAL:
        return resolve_special_class(master)
    raise ClassificationError("INVALID_REGISTRATION_ROUTE", "登録経路が不正です。")


def assigned_to_machine_class(master, machine_class):
    if master is None:
        return False
    return MachineAssignment.objects.filter(
        code=master, machine__machine_class=str(machine_class)
    ).exists()


def sync_master_class_from_assignment(master):
    if master is None:
        return
    classes = _assignment_classes(master)
    if classes & {1, 2} == {1, 2}:
        raise ClassificationError(
            "CLASS_1_2_CONFLICT",
            "クラス1と2の機械設定が重複しています。機械設定を確認してください。",
            {
                "code": master.code,
                "detected_classes": [1, 2],
                "machine_numbers": _conflicting_machine_numbers(master),
            },
        )
    MasterClass.objects.filter(master=master, class_master__class_no__in=[1, 2]).delete()
    ac = next(iter(classes & {1, 2}), None)
    if ac is not None and ac in (1, 2):
        cm = ClassMaster.objects.filter(class_no=ac).first()
        if cm:
            MasterClass.objects.update_or_create(
                master=master,
                class_master=cm,
                defaults={"class_master": cm},
            )
    if ac is None and assigned_to_machine_class(master, 3):
        cm = ClassMaster.objects.filter(class_no=3).first()
        if cm:
            MasterClass.objects.update_or_create(
                master=master,
                class_master=cm,
                defaults={"class_master": cm},
            )


def determine_inspection_class(master, file_codes_by_code):
    if master is None:
        return None
    code = master.code

    ac = get_assignment_class_for_master(master)
    if ac is not None and ac in (1, 2):
        return ac

    if assigned_to_machine_class(master, 3):
        return 3

    n1 = (master.node_type_1 or "").strip()
    dept = (master.department or "").strip()

    file_infos = file_codes_by_code.get(code, [])
    folder_classes = set()
    for fi in file_infos:
        fc = classify_folder(fi["file_path"])
        if fc:
            folder_classes.add(fc)

    if "product1" in folder_classes:
        return 6
    if "product2" in folder_classes:
        return 7

    if n1 == "プレス":
        return 4
    if n1 == "加工" and (dept == "製造管理部" or dept == "生残技術部"):
        return 5

    return 8


def get_ocr_engine():
    global _OCR_ENGINE
    if _OCR_ENGINE is None:
        _OCR_ENGINE = RapidOCR()
    return _OCR_ENGINE


@transaction.atomic
def upsert_targets(target_date, codes, registration_route, user=None):
    session, _ = InspectionSession.objects.get_or_create(
        target_date=target_date,
        owner_user=user,
        defaults={"created_by": user, "updated_by": user},
    )
    added_count = 0
    duplicate_count = 0

    for raw_code in codes:
        normalized_code = normalize_code(raw_code)
        if not normalized_code:
            continue
        master = Master.objects.filter(code=normalized_code).first()

        if master is None and registration_route == InspectionTarget.RegistrationRoute.SPECIAL:
            resolve_special_class(master)
        resolved_class = None if master is None else resolve_class_for_route(master, registration_route)
        lookup_kwargs = {
            "session": session,
            "normalized_code": normalized_code,
            "class_override": resolved_class,
        }
        requires_sheet = inspection_sheet_required_for_class(resolved_class)
        issue_status = (
            InspectionTarget.IssueStatus.PENDING
            if requires_sheet
            else InspectionTarget.IssueStatus.NOT_REQUIRED
        )

        target, created = InspectionTarget.objects.get_or_create(
            **lookup_kwargs,
            defaults={
                "master": master,
                "raw_code": raw_code,
                "class_override": resolved_class,
                "registration_route": registration_route,
                "created_by": user,
                "updated_by": user,
                "source_ocr": registration_route == "ocr",
                "source_excel": registration_route == "excel",
                "source_manual": registration_route in ("manual_code", "factory_map", "special"),
                "requires_inspection_sheet": requires_sheet,
                "issue_status": issue_status,
            },
        )

        if not created:
            duplicate_count += 1
            if registration_route == "ocr":
                target.source_ocr = True
            elif registration_route == "excel":
                target.source_excel = True
            else:
                target.source_manual = True
            if master and not target.master:
                target.master = master
            target.requires_inspection_sheet = requires_sheet
            if requires_sheet and target.issue_status == InspectionTarget.IssueStatus.NOT_REQUIRED:
                target.issue_status = InspectionTarget.IssueStatus.PENDING
            target.save(
                update_fields=[
                    "source_manual",
                    "source_ocr",
                    "source_excel",
                    "master",
                    "requires_inspection_sheet",
                    "issue_status",
                    "updated_at",
                ]
            )
        else:
            added_count += 1

        if master is None:
            InspectionTargetWarning.objects.get_or_create(
                target=target,
                error_code="UNKNOWN_CODE",
                defaults={
                    "message": f"Unknown code: {normalized_code}",
                    "details": {"code": normalized_code},
                },
            )

    return session, added_count, duplicate_count


def add_manual_targets(target_date, codes, user=None):
    session, added_count, _ = upsert_targets(target_date, codes, "manual_code", user=user)
    return session, added_count


def add_factory_map_target(target_date, machine_id, code, user=None):
    master = Master.objects.filter(code=normalize_code(code)).first()
    if master is None or not MachineAssignment.objects.filter(machine_id=machine_id, code=master).exists():
        raise ClassificationError("INVALID_REQUEST", "指定した機械に品目が割り当てられていません。", {"code": normalize_code(code)})
    session, added_count, _ = upsert_targets(target_date, [code], "factory_map", user=user)
    return session, added_count


def add_special_targets(target_date, codes, user=None):
    session, added_count, _ = upsert_targets(target_date, codes, "special", user=user)
    return session, added_count


def extract_codes_from_text(text):
    return [normalize_code(match.group(0)) for match in CODE_PATTERN.finditer(text or "")]


def find_code_candidates(text):
    return [normalize_code(match.group(0)) for match in CODE_PATTERN.finditer(text or "")]


def extract_codes_from_uploaded_text(uploaded_file):
    raw = uploaded_file.read()
    for encoding in ("utf-8-sig", "cp932", "utf-8"):
        try:
            return find_code_candidates(raw.decode(encoding))
        except UnicodeDecodeError:
            continue
    return find_code_candidates(raw.decode("utf-8", errors="ignore"))


def extract_codes_from_plan_excel(path_or_file, sheet_name, file_name=None):
    name = file_name or str(path_or_file)
    suffix = Path(name).suffix.lower()
    values = []
    if suffix == ".xls":
        if hasattr(path_or_file, "read"):
            workbook = xlrd.open_workbook(file_contents=path_or_file.read())
        else:
            workbook = xlrd.open_workbook(str(path_or_file))
        worksheet = workbook.sheet_by_name(sheet_name)
        for row_index in range(3, 103):
            val = worksheet.cell_value(row_index, 5)
            if val:
                values.append(val)
    else:
        workbook = load_workbook(path_or_file, read_only=True, data_only=True)
        worksheet = workbook[sheet_name]
        for row in range(4, 104):
            val = worksheet.cell(row=row, column=6).value
            if val:
                values.append(val)

    codes = []
    for value in values:
        codes.extend(find_code_candidates(str(value)))
    return codes


def extract_codes_and_match_failures_from_pdf(scan_file):
    temp_dir = Path(settings.TEST_INPUT_DIR)
    temp_dir.mkdir(parents=True, exist_ok=True)
    temp_pdf_path = temp_dir / f"ocr_{uuid4().hex}.pdf"
    temp_pdf_path.write_bytes(scan_file.read())

    embedded_text = ""
    try:
        reader = PdfReader(str(temp_pdf_path))
        embedded_text = "\n".join((page.extract_text() or "") for page in reader.pages)
    except Exception as exc:
        logger.warning("Embedded PDF text extraction failed: %s", exc)

    if embedded_text.strip():
        codes = find_code_candidates(embedded_text)
        match_failed_count = 0 if codes else 1
        temp_pdf_path.unlink(missing_ok=True)
        return {
            "codes": codes,
            "match_failed_count": match_failed_count,
            "page_count": len(reader.pages),
            "mode": "embedded_text",
        }

    pdf = fitz.open(str(temp_pdf_path))
    page_texts = []
    for page_index in range(pdf.page_count):
        pixmap = pdf[page_index].get_pixmap(matrix=fitz.Matrix(3, 3), alpha=False)
        image_path = temp_dir / f"{temp_pdf_path.stem}_page_{page_index + 1}.png"
        pixmap.save(str(image_path))
        result, _ = get_ocr_engine()(str(image_path))
        lines = [line[1] for line in result] if result else []
        page_texts.append("\n".join(lines))
        image_path.unlink(missing_ok=True)

    pdf.close()
    temp_pdf_path.unlink(missing_ok=True)

    codes = []
    match_failed_count = 0
    for text in page_texts:
        page_codes = find_code_candidates(text)
        codes.extend(page_codes)
        if not page_codes and ALNUM_PATTERN.search(text or ""):
            match_failed_count += 1

    return {
        "codes": codes,
        "match_failed_count": match_failed_count,
        "page_count": len(page_texts),
        "mode": "ocr",
    }


def _read_csv_text(path_or_file):
    if hasattr(path_or_file, "read"):
        raw = path_or_file.read()
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw.decode("cp932", errors="replace")
    else:
        for enc in ("utf-8-sig", "cp932"):
            try:
                text = Path(path_or_file).read_text(encoding=enc)
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
        else:
            text = Path(path_or_file).read_text(encoding="cp932", errors="replace")
    return text


def load_master_rows_from_csv(path_or_file):
    text = _read_csv_text(path_or_file)
    rows = []
    reader = csv.reader(text.splitlines())
    header = next(reader, None)
    for row in reader:
        if len(row) < 17:
            continue
        code = normalize_code(row[9])
        name = str(row[10]).strip()
        department = str(row[16]).strip() if len(row) > 16 else ""
        root_code = normalize_code(row[1]) if len(row) > 1 else ""
        parent_code = normalize_code(row[8]) if len(row) > 8 else ""
        level_str = str(row[11]).strip() if len(row) > 11 else ""
        quantity_str = str(row[18]).strip() if len(row) > 18 else ""
        if not code or not name:
            continue
        try:
            level = int(level_str) if level_str else 0
        except ValueError:
            level = 0
        try:
            quantity = float(quantity_str) if quantity_str else 0
        except ValueError:
            quantity = 0
        rows.append(
            {
                "code": code,
                "name": name,
                "department": department,
                "root_code": root_code,
                "parent_code": parent_code,
                "level": level,
                "quantity": quantity,
            }
        )
    return rows


def setter_check(folders):
    return "setter" in folders


def product_inspection_1(folders):
    return "product1" in folders


def product_inspection_2(folders):
    return "product2" in folders


def scan_inspection_folders(folder_paths):
    file_map = {}
    warnings = []
    for path_str in folder_paths:
        path_str = path_str.strip()
        if not path_str:
            continue
        p = Path(path_str)
        if not p.exists() or not p.is_dir():
            warnings.append(f"Folder not found or inaccessible: {path_str}")
            continue
        for f in p.rglob("*"):
            if f.is_file():
                name = f.name
                matches = list(CODE_PATTERN.finditer(name.upper()))
                for m in matches:
                    code = normalize_code(m.group(0))
                    file_map.setdefault(code, []).append({
                        "file_name": name,
                        "file_path": str(f.resolve()),
                    })
    return file_map, warnings


def classify_folder(file_path):
    lower = file_path.lower()
    if "自動機" in lower and "工程内検査" in lower:
        return "auto1"
    if "巡回検査" in lower and "自動機" in lower:
        return "auto2"
    if "巡回検査" in lower and "セッター" in lower:
        return "setter"
    if "製品検査" in lower and "(1)" in lower:
        return "product1"
    if "製品検査" in lower and "(2)" in lower:
        return "product2"
    return None


def resolve_inspection_file(master, class_no):
    if master is None:
        return None
    if class_no == 9:
        setting = SpecialInspectionClass9.objects.filter(master=master).first()
        if not setting or not setting.inspection_sheet_path:
            return None
        return {"file_path": setting.inspection_sheet_path, "file_name": Path(setting.inspection_sheet_path).name}
    wanted_folder = {1: "auto1", 6: "product1", 7: "product2"}.get(class_no)
    if wanted_folder is None:
        return None
    matches = [
        row
        for row in InspectionFile.objects.filter(master=master).order_by("id")
        if classify_folder(row.file_path) == wanted_folder
    ]
    return select_preferred_inspection_file(matches)


def _normalized_file_path(file_path):
    return os.path.normpath(str(file_path)).replace("\\", "/").casefold()


def select_preferred_inspection_file(inspection_files):
    return min(
        inspection_files,
        key=lambda row: (
            -row.priority,
            0 if row.file_created is not None else 1,
            -row.file_created.timestamp() if row.file_created is not None else 0,
            _normalized_file_path(row.file_path),
            row.id,
        ),
        default=None,
    )


def resolve_unambiguous_inspection_file(master):
    candidate_classes = set()
    inspection_files = list(InspectionFile.objects.filter(master=master).order_by("id"))
    for file_path in (row.file_path for row in inspection_files):
        folder_class = classify_folder(file_path)
        class_no = {"auto1": 1, "product1": 6, "product2": 7}.get(folder_class)
        if class_no:
            candidate_classes.add(class_no)
    if SpecialInspectionClass9.objects.filter(master=master).exclude(inspection_sheet_path="").exists():
        candidate_classes.add(9)
    if len(candidate_classes) > 1:
        raise ClassificationError(
            "AMBIGUOUS_INSPECTION_FILE",
            "複数クラスの検査書があります。対象一覧から検査を選択してください。",
            {"code": master.code, "detected_classes": sorted(candidate_classes)},
        )
    if candidate_classes:
        return resolve_inspection_file(master, next(iter(candidate_classes)))
    return select_preferred_inspection_file(inspection_files)


def _file_created_at(path):
    try:
        file_stat = path.stat()
        if hasattr(file_stat, "st_birthtime"):
            timestamp = file_stat.st_birthtime
        elif os.name == "nt":
            timestamp = file_stat.st_ctime
        else:
            return None
        return datetime.fromtimestamp(timestamp, tz=timezone.get_current_timezone())
    except (OSError, OverflowError, ValueError):
        return None


def scan_and_classify_files(folder_paths, folder_priorities=None):
    code_file_map = {}
    folder_warnings = []
    priorities = folder_priorities or {}
    for path_str in folder_paths:
        path_str = path_str.strip()
        if not path_str:
            continue
        p = Path(path_str)
        try:
            accessible = p.exists() and p.is_dir()
        except OSError:
            accessible = False
        if not accessible:
            folder_warnings.append(f"Folder not found or inaccessible: {path_str}")
            continue
        try:
            for f in p.rglob("*"):
                if f.is_file():
                    name = f.name
                    fpath = str(f.resolve())
                    for m in FILE_CODE_PATTERN.finditer(name.upper()):
                        code = normalize_code(m.group(0))
                        code_file_map.setdefault(code, []).append({
                            "file_name": name,
                            "file_path": fpath,
                            "priority": priorities.get(path_str, 0),
                            "file_created": _file_created_at(f),
                        })
        except OSError:
            raise OSError(INSPECTION_FOLDER_SCAN_ERROR) from None
    return code_file_map, folder_warnings


@transaction.atomic
def import_master_csv(
    master_file=None,
    csv_path=None,
    inspection_folder_paths=None,
    inspection_folder_priorities=None,
):
    update_started = perf_counter()
    if master_file:
        rows = load_master_rows_from_csv(master_file)
        source = getattr(master_file, "name", "uploaded")
    elif csv_path:
        rows = load_master_rows_from_csv(csv_path)
        source = csv_path
    else:
        source_path = Path(settings.TEST_INPUT_DIR) / "master.csv"
        rows = load_master_rows_from_csv(source_path)
        source = str(source_path)

    folder_paths = inspection_folder_paths or []
    code_file_map, folder_warnings = scan_and_classify_files(folder_paths, inspection_folder_priorities)
    configured_folder_count = len([path for path in folder_paths if str(path).strip()])
    all_configured_folders_inaccessible = (
        configured_folder_count > 0
        and not code_file_map
        and len(folder_warnings) >= configured_folder_count
    )
    if folder_warnings and not all_configured_folders_inaccessible:
        raise OSError(INSPECTION_FOLDER_SCAN_ERROR)

    processed_codes = set()
    master_count = 0
    class_count = 0
    structure_count = 0

    for row in rows:
        code = row["code"]
        if code in processed_codes:
            continue
        processed_codes.add(code)

        classification = classify_master_by_product_code(code)
        master, created = Master.objects.update_or_create(
            code=code,
            defaults={
                "name": row["name"],
                "node_type_1": classification["node_type_1"] or None,
                "node_type_2": classification["node_type_2"] or None,
                "product_category": classification["category"] or None,
                "department": row["department"],
            },
        )
        master_count += 1

        resolved_classes = []
        try:
            resolved_classes.append(resolve_process_class(master))
        except ClassificationError as exc:
            if exc.error_code != "PROCESS_CLASS_NOT_FOUND":
                raise

        file_folder_classes = {
            classify_folder(info["file_path"])
            for info in code_file_map.get(code, [])
        } & {"product1", "product2"}
        if file_folder_classes == {"product1", "product2"}:
            candidate_file_names = [
                _safe_file_name(info.get("file_name") or info["file_path"])
                for info in code_file_map.get(code, [])
                if classify_folder(info["file_path"]) in file_folder_classes
            ]
            raise ClassificationError(
                "CLASS_6_7_CONFLICT",
                "製品検査(1)と製品検査(2)の両方に検査書があります。保存場所を確認してください。",
                {
                    "code": code,
                    "detected_classes": [6, 7],
                    "candidate_file_names": candidate_file_names,
                },
            )
        if "product1" in file_folder_classes:
            resolved_classes.append(6)
        elif "product2" in file_folder_classes:
            resolved_classes.append(7)

        for insp_class in resolved_classes:
            cm = ClassMaster.objects.filter(class_no=insp_class).first()
            if cm is None:
                continue
            MasterClass.objects.update_or_create(
                master=master,
                class_master=cm,
                defaults={"class_master": cm},
            )
            class_count += 1

    for row in rows:
        parent_code = row["parent_code"]
        child_code = row["code"]
        if not parent_code or not child_code:
            continue
        Structure.objects.update_or_create(
            parent_code=parent_code,
            child_code=child_code,
            defaults={
                "root_code": row["root_code"],
                "level": row["level"],
                "quantity": row["quantity"] if row["quantity"] != 0 else None,
            },
        )
        structure_count += 1

    if all_configured_folders_inaccessible:
        file_count = InspectionFile.objects.count()
    else:
        InspectionFile.objects.all().delete()
        file_count = 0
        for code, file_infos in code_file_map.items():
            master = Master.objects.filter(code=code).first()
            if master is None:
                continue
            for fi in file_infos:
                InspectionFile.objects.create(
                    master=master,
                    file_name=fi["file_name"],
                    file_path=fi["file_path"],
                    priority=fi["priority"],
                    file_created=fi["file_created"],
                )
                file_count += 1

    result = {
        "updated_master_count": master_count,
        "updated_class_count": class_count,
        "updated_structure_count": structure_count,
        "inspection_file_count": file_count,
        "source": source,
        "total_update_seconds": round(perf_counter() - update_started, 3),
        "transaction_strategy": "single_atomic_update",
    }
    if folder_warnings:
        result["folder_warnings"] = folder_warnings
    if all_configured_folders_inaccessible:
        result["inspection_files_preserved"] = True
    return result


@transaction.atomic
def import_plan_targets(target_date, scan_file=None, excel_file=None, sheet_name=None, user=None):
    sources = []
    warning_summary = {
        "UNKNOWN_CODE": 0,
        "DUPLICATE_TARGET": 0,
        "MATCH_FAILED": 0,
    }

    if scan_file:
        scan_name = getattr(scan_file, "name", "").lower()
        if scan_name.endswith(".pdf"):
            ocr_result = extract_codes_and_match_failures_from_pdf(scan_file)
            ocr_codes = ocr_result["codes"]
            warning_summary["MATCH_FAILED"] += ocr_result["match_failed_count"]
            source_details = {
                "source": "ocr",
                "read_count": len(ocr_codes),
                "match_failed_count": ocr_result["match_failed_count"],
                "page_count": ocr_result["page_count"],
                "mode": ocr_result["mode"],
            }
        else:
            ocr_codes = extract_codes_from_uploaded_text(scan_file)
            source_details = {
                "source": "ocr",
                "read_count": len(ocr_codes),
                "match_failed_count": 0,
                "mode": "text",
            }

        _, added_count, duplicate_count = upsert_targets(target_date, ocr_codes, "ocr", user=user)
        warning_summary["DUPLICATE_TARGET"] += duplicate_count
        sources.append(
            {
                "added_count": added_count,
                "duplicate_count": duplicate_count,
                **source_details,
            }
        )

    if excel_file:
        excel_codes = extract_codes_from_plan_excel(excel_file, sheet_name, excel_file.name)
        _, added_count, duplicate_count = upsert_targets(target_date, excel_codes, "excel", user=user)
        warning_summary["DUPLICATE_TARGET"] += duplicate_count
        sources.append(
            {
                "source": "excel",
                "read_count": len(excel_codes),
                "added_count": added_count,
                "duplicate_count": duplicate_count,
            }
        )

    session, _ = InspectionSession.objects.get_or_create(
        target_date=target_date, owner_user=user,
        defaults={"created_by": user, "updated_by": user},
    )
    warning_count = InspectionTargetWarning.objects.filter(target__session=session).count()
    warning_summary["UNKNOWN_CODE"] = warning_count
    imported_count = InspectionTarget.objects.filter(session=session).count()
    result = {
        "target_date": str(target_date),
        "session_id": session.id,
        "imported_count": imported_count,
        "warning_count": warning_count,
        "warning_summary": warning_summary,
        "sources": sources,
    }
    return result


def history_map_for_date(target_date, user=None):
    rows = History.objects.filter(date=target_date, created_by=user, deleted_at__isnull=True).values_list("master_id", "time_slot", "class_override")
    history_map = {}
    for master_id, time_slot, co in rows:
        history_map.setdefault((master_id, co), set()).add(time_slot)
    return history_map


@transaction.atomic
def set_check(target_date, target_id, time_slot, checked, user=None):
    target = InspectionTarget.objects.select_related("master", "session").filter(
        id=target_id, session__target_date=target_date, session__owner_user=user, deleted_at__isnull=True
    ).first()
    if target is None or target.master is None:
        raise ClassificationError("TARGET_NOT_FOUND", "検査対象が見つかりません。")
    master = target.master
    class_override = target.class_override

    lookup = {
        "created_by": user,
        "date": target_date,
        "master": master,
        "time_slot": time_slot,
    }
    if class_override:
        lookup["class_override"] = class_override
    else:
        lookup["class_override__isnull"] = True

    history = History.objects.filter(**lookup).first()
    if checked:
        if history:
            history.deleted_at = None
            history.deleted_by = None
            history.updated_by = user
            history.save(update_fields=["deleted_at", "deleted_by", "updated_by", "updated_at"])
        else:
            create_data = {key: value for key, value in lookup.items() if not key.endswith("__isnull")}
            create_data.setdefault("class_override", class_override)
            History.objects.create(**create_data, updated_by=user)
    elif history and history.deleted_at is None:
        history.deleted_at = timezone.now()
        history.deleted_by = user
        history.updated_by = user
        history.save(update_fields=["deleted_at", "deleted_by", "updated_by", "updated_at"])


@transaction.atomic
def bulk_upsert_history(target_date, items, user=None):
    updated = 0
    for item in items:
        for slot, checked in item["checks"].items():
            set_check(target_date, item["target_id"], slot, checked, user=user)
        updated += 1
    return updated


def create_job(job_type, payload, user=None):
    timestamp = timezone.localtime().strftime("%Y%m%d%H%M%S")
    return Job.objects.create(
        job_id=f"job_{timestamp}_{uuid4().hex[:8]}",
        job_type=job_type,
        request_payload=payload,
        created_by=user,
        updated_by=user,
    )


class StaleJobExecution(RuntimeError):
    pass


def run_job(
    job,
    fn,
    mark_running=True,
    *,
    worker_id=None,
    execution_token=None,
):
    if mark_running:
        job.status = Job.Status.RUNNING
        job.started_at = timezone.now()
        job.updated_by = job.created_by
        job.save(update_fields=["status", "started_at", "updated_by", "updated_at"])
    execution_started = perf_counter()
    try:
        if execution_token:
            with transaction.atomic():
                _lock_id = make_advisory_lock_id(job.job_id, execution_token)
                _app_name = make_application_name_marker(job.job_id, execution_token)
                with connection.cursor() as cursor:
                    cursor.execute("SELECT pg_advisory_xact_lock(%s, %s)",
                                   [ADVISORY_LOCK_NAMESPACE, _lock_id])
                    cursor.execute("SET LOCAL application_name = %s", [_app_name])
                result = fn()
        else:
            result = fn()
    except Exception as exc:
        error_message = str(exc)
        failure_result = {
            "status": "failed",
            "error_message": error_message,
            "exception_type": type(exc).__name__,
        }
        if isinstance(exc, ClassificationError):
            safe_details = _safe_classification_details(exc.details)
            failure_result["error_code"] = exc.error_code
            failure_result["details"] = safe_details
            error_message = _classification_job_error_message(error_message, safe_details)
        finished_at = timezone.now()
        if execution_token:
            updated = Job.objects.filter(
                pk=job.pk,
                status=Job.Status.RUNNING,
                worker_id=worker_id,
                execution_token=execution_token,
            ).update(
                status=Job.Status.FAILED,
                error_message=error_message,
                result=failure_result,
                finished_at=finished_at,
                updated_by=job.created_by,
                worker_id="",
                execution_token="",
                heartbeat_at=None,
                lease_until=None,
                updated_at=finished_at,
            )
            if updated != 1:
                raise StaleJobExecution("Job execution lease is no longer owned.") from exc
            return failure_result
        else:
            job.status = Job.Status.FAILED
            job.error_message = error_message
            job.result = failure_result
            job.finished_at = finished_at
            job.updated_by = job.created_by
            job.worker_id = ""
            job.execution_token = ""
            job.heartbeat_at = None
            job.lease_until = None
            job.save(update_fields=["status", "error_message", "result", "finished_at", "updated_by", "worker_id", "execution_token", "heartbeat_at", "lease_until", "updated_at"])
        raise

    if isinstance(result, dict):
        result.setdefault(
            "job_metrics",
            {
                "execution_seconds": round(perf_counter() - execution_started, 3),
                "attempt_count": job.attempt_count,
            },
        )
    finished_at = timezone.now()
    if execution_token:
        updated = Job.objects.filter(
            pk=job.pk,
            status=Job.Status.RUNNING,
            worker_id=worker_id,
            execution_token=execution_token,
        ).update(
            status=Job.Status.SUCCEEDED,
            result=result,
            finished_at=finished_at,
            updated_by=job.created_by,
            worker_id="",
            execution_token="",
            heartbeat_at=None,
            lease_until=None,
            updated_at=finished_at,
        )
        if updated != 1:
            raise StaleJobExecution("Job execution lease is no longer owned.")
    else:
        job.status = Job.Status.SUCCEEDED
        job.result = result
        job.finished_at = finished_at
        job.updated_by = job.created_by
        job.worker_id = ""
        job.execution_token = ""
        job.heartbeat_at = None
        job.lease_until = None
        job.save(update_fields=["status", "result", "finished_at", "updated_by", "worker_id", "execution_token", "heartbeat_at", "lease_until", "updated_at"])
    return result


def run_erp_automation(erp_path, csv_path):
    module_dir = Path(__file__).resolve().parent.parent.parent / "erp_automation"
    script = module_dir / "erp.py"
    if not script.exists():
        raise FileNotFoundError(f"スクリプトが見つかりません: {script}")
    try:
        result = subprocess.run(
            [sys.executable, str(script), erp_path, csv_path],
            capture_output=True,
            text=True,
            timeout=600,
        )
    except subprocess.TimeoutExpired as exc:
        raise RuntimeError("ERP automation timed out (600s)") from exc
    if result.returncode != 0:
        raise RuntimeError(f"ERP automation failed: {result.stderr.strip()}")
    return {"status": "succeeded", "output": result.stdout.strip()}


def _safe_classification_details(details):
    details = details if isinstance(details, dict) else {}
    safe_details = {}
    if isinstance(details.get("code"), str):
        safe_details["code"] = details["code"]
    if isinstance(details.get("class"), int):
        safe_details["class"] = details["class"]
    if isinstance(details.get("candidate_count"), int):
        safe_details["candidate_count"] = details["candidate_count"]
    for key in ("detected_classes",):
        values = details.get(key)
        if isinstance(values, (list, tuple)):
            safe_details[key] = [value for value in values if isinstance(value, int)]
    candidate_file_names = details.get("candidate_file_names")
    if isinstance(candidate_file_names, (list, tuple)):
        safe_details["candidate_file_names"] = [
            _safe_file_name(value) for value in candidate_file_names if isinstance(value, str)
        ]
    machine_numbers = details.get("machine_numbers")
    if isinstance(machine_numbers, (list, tuple)):
        safe_details["machine_numbers"] = [value for value in machine_numbers if isinstance(value, str)]
    return safe_details


def _classification_job_error_message(message, details):
    parts = []
    if details.get("code"):
        parts.append(f"品番: {details['code']}")
    if details.get("class") is not None:
        parts.append(f"class: {details['class']}")
    elif details.get("detected_classes"):
        parts.append("class: " + "/".join(str(value) for value in details["detected_classes"]))
    if details.get("candidate_count") is not None:
        parts.append(f"候補: {details['candidate_count']}件")
    if details.get("candidate_file_names"):
        parts.append("検査書: " + ", ".join(details["candidate_file_names"]))
    if details.get("machine_numbers"):
        parts.append("機械: " + ", ".join(details["machine_numbers"]))
    return f"{message} {' / '.join(parts)}" if parts else message


def _period_for_date(d: date_type) -> int:
    if d.month >= 10:
        return d.year - 1956
    return d.year - 1 - 1956


def _time_slot_to_am_pm(slot: str) -> str:
    return "AM" if slot in ("A", "B") else "PM"


def write_history_to_excel(target_date: date_type, user=None) -> int:
    from .models import AppSetting, InspectionSession

    setting = AppSetting.objects.first()
    if not setting or not setting.history_file_path:
        raise FileNotFoundError("履歴ファイルのパスが設定されていません。設定タブで履歴ファイルのパスを指定してください。")

    history_path = Path(setting.history_file_path)
    if not history_path.exists():
        raise FileNotFoundError("履歴ファイルが保存場所に存在しません。")

    master_ids_class_1 = MasterClass.objects.filter(
        class_master__class_no=1
    ).values_list("master_id", flat=True)

    qs = (
        History.objects.filter(date=target_date, created_by=user, deleted_at__isnull=True)
        .filter(Q(class_override=1) | Q(class_override__isnull=True, master_id__in=master_ids_class_1))
        .select_related("master")
        .order_by("master__code", "time_slot")
    )
    if not qs.exists():
        return 0

    period = _period_for_date(target_date)
    sheet_name = f"{period}期"

    import win32com.client
    import pythoncom

    pythoncom.CoInitialize()
    xl = None
    try:
        xl = win32com.client.DispatchEx("Excel.Application")
        xl.DisplayAlerts = False
        xl.EnableEvents = False

        wb = xl.Workbooks.Open(str(history_path.resolve()))

        ws = None
        for s in wb.Sheets:
            if s.Name == sheet_name:
                ws = s
                break
        if ws is None:
            ws = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
            ws.Name = sheet_name

        existing_keys = set()
        first_blank_row = None
        try:
            max_row = ws.UsedRange.Rows.Count
        except Exception:
            max_row = 1

        for row_idx in range(2, max(max_row, 1) + 2):
            b_val = ws.Cells(row_idx, 2).Value
            if b_val is None and first_blank_row is None:
                first_blank_row = row_idx
                break
            c_val = ws.Cells(row_idx, 3).Value
            d_val = ws.Cells(row_idx, 4).Value
            if b_val is not None and c_val is not None and d_val is not None:
                existing_keys.add((b_val, str(c_val).strip(), str(d_val).strip()))

        if first_blank_row is None:
            first_blank_row = max(max_row, 1) + 1

        next_row = first_blank_row

        written_count = 0
        for h in qs:
            code = h.master.code if h.master else ""
            name = h.master.name if h.master else ""
            serial_date = _excel_serial_date(h.date)
            am_pm = _time_slot_to_am_pm(h.time_slot)
            key = (serial_date, am_pm, code)
            if key in existing_keys:
                continue
            existing_keys.add(key)

            ws.Cells(next_row, 2).Value = serial_date
            ws.Cells(next_row, 3).Value = am_pm
            ws.Cells(next_row, 4).Value = code
            ws.Cells(next_row, 5).Value = name
            ws.Cells(next_row, 6).Value = "合格"
            next_row += 1
            written_count += 1

        wb.Save()
        xl.Visible = True
    finally:
        if xl is not None:
            try:
                xl.EnableEvents = True
            except pythoncom.com_error:
                pass
        pythoncom.CoUninitialize()

    session = InspectionSession.objects.filter(target_date=target_date, owner_user=user).first()
    if session:
        session.history = True
        session.save(update_fields=["history", "updated_at"])

    return written_count


def _excel_serial_date(d: date_type) -> int:
    delta = d - date_type(1899, 12, 30)
    return delta.days


def convert_excel_to_pdf(file_path: str) -> str:
    import os as os_mod
    import pythoncom
    import win32com.client

    if not os_mod.path.exists(file_path):
        raise FileNotFoundError("検査書ファイルが保存場所に存在しません。")

    ext = os_mod.path.splitext(file_path)[1].lower()
    if ext not in (".xls", ".xlsx", ".xlsm"):
        raise ValueError(f"Excelファイルではありません: {file_path}")

    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_file:
        tmp_path = tmp_file.name
    pythoncom.CoInitialize()
    xl = None
    try:
        xl = win32com.client.DispatchEx("Excel.Application")
        xl.DisplayAlerts = False
        xl.Visible = False
        wb = xl.Workbooks.Open(file_path)
        try:
            wb.ExportAsFixedFormat(0, tmp_path)
        finally:
            wb.Close(False)
    except Exception:
        try:
            os_mod.unlink(tmp_path)
        except OSError:
            pass
        raise
    finally:
        if xl:
            xl.Quit()
        pythoncom.CoUninitialize()
    return tmp_path


def _print_file_direct(file_path: str):
    import os as os_mod
    import pythoncom
    import win32com.client
    import win32api

    if not os_mod.path.exists(file_path):
        raise FileNotFoundError("検査書ファイルが保存場所に存在しません。")

    ext = os_mod.path.splitext(file_path)[1].lower()

    pythoncom.CoInitialize()
    try:
        if ext in (".xls", ".xlsx", ".xlsm"):
            xl = None
            try:
                xl = win32com.client.DispatchEx("Excel.Application")
                xl.DisplayAlerts = False
                xl.Visible = False
                wb = xl.Workbooks.Open(file_path)
                for ws in wb.Sheets:
                    ws.PageSetup.BlackAndWhite = True
                wb.PrintOut()
                wb.Close(False)
            finally:
                if xl:
                    xl.Quit()
        elif ext == ".pdf":
            import fitz
            import tempfile
            doc = fitz.open(file_path)
            new_doc = fitz.open()
            tmp_path = None
            try:
                for page in doc:
                    rect = page.rect
                    pix = page.get_pixmap(dpi=200, colorspace=fitz.csGRAY)
                    new_page = new_doc.new_page(width=rect.width, height=rect.height)
                    new_page.insert_image(rect, stream=pix.tobytes("png"))
                tmp_path = tempfile.mktemp(suffix=".pdf")
                new_doc.save(tmp_path)
                win32api.ShellExecute(0, "print", tmp_path, None, ".", 0)
            finally:
                doc.close()
                new_doc.close()
                if tmp_path and os_mod.path.exists(tmp_path):
                    try:
                        os_mod.unlink(tmp_path)
                    except OSError:
                        pass
        else:
            win32api.ShellExecute(0, "print", file_path, None, ".", 0)
    finally:
        pythoncom.CoUninitialize()


def issue_inspection_sheets(target_date: date_type | None = None, user=None):
    template_path = Path(settings.DAILY_REPORT_TEMPLATE)
    if not template_path.exists():
        raise FileNotFoundError("検査書テンプレートが保存場所に存在しません。")

    qs = History.objects.filter(is_sheet_issued=False, created_by=user, deleted_at__isnull=True)
    if target_date:
        qs = qs.filter(date=target_date)
    qs = qs.select_related("master").order_by("date", "master__code", "time_slot")

    if not qs.exists():
        return {"issued_count": 0, "message": "No unissued entries found."}

    legacy_class_map = {}
    for master_id, class_no in MasterClass.objects.filter(
        master_id__in=qs.values_list("master_id", flat=True), class_master__isnull=False
    ).exclude(class_master__class_no=9).values_list("master_id", "class_master__class_no"):
        legacy_class_map.setdefault(master_id, class_no)

    rows = []
    target_history_ids: list[int] = []
    class9_history_ids: list[int] = []
    for h in qs:
        class_no = h.class_override or legacy_class_map.get(h.master_id)
        if class_no == 9:
            class9_history_ids.append(h.history_id)
            continue
        if class_no not in (1, 6, 7):
            continue
        resolved_file = resolve_inspection_file(h.master, class_no)
        if resolved_file is None and h.class_override is None:
            resolved_file = resolve_unambiguous_inspection_file(h.master)
        if not resolved_file:
            continue
        serial_date = _excel_serial_date(h.date)
        rows.append([
            resolved_file.file_path,
            h.time_slot,
            class_no if class_no is not None else "",
            serial_date,
            "1-6",
            "前田 賢一",
        ])
        target_history_ids.append(h.history_id)

    class9_successful_ids: list[int] = []
    class9_warnings: list[dict] = []
    if class9_history_ids:
        for h in qs.filter(history_id__in=class9_history_ids).select_related("master"):
            resolved_file = resolve_inspection_file(h.master, 9)
            if not resolved_file:
                class9_warnings.append({
                    "history_id": h.history_id,
                    "code": h.master.code if h.master else "",
                    "error_code": "FILE_NOT_FOUND",
                    "message": "特殊検査の検査書ファイルが設定されていません。",
                })
                continue
            try:
                _print_file_direct(resolved_file["file_path"])
                class9_successful_ids.append(h.history_id)
            except Exception:
                logger.exception("Class 9 print failed for history_id=%s", h.history_id)
                class9_warnings.append({
                    "history_id": h.history_id,
                    "code": h.master.code if h.master else "",
                    "error_code": "PRINT_FAILED",
                    "message": "特殊検査の印刷に失敗しました。",
                })

    if class9_successful_ids:
        History.objects.filter(history_id__in=class9_successful_ids).update(
            is_sheet_issued=True, updated_by=user, updated_at=timezone.now()
        )

    if not rows and not class9_successful_ids:
        return {
            "issued_count": 0,
            "class9_printed": 0,
            "class9_failed": len(class9_warnings),
            "warnings": class9_warnings,
            "message": "No printable entries found (no valid file paths or non-target classes).",
        }

    output_path = None
    if rows:
        output_dir = Path(settings.DAILY_REPORT_OUTPUT_DIR)
        output_dir.mkdir(parents=True, exist_ok=True)
        owner_suffix = f"u{user.pk}" if user is not None else "legacy"
        date_suffix = str(target_date).replace("-", "") if target_date else "all"
        output_path = output_dir / f"inspection_sheets_{date_suffix}_{owner_suffix}.xlsm"
        shutil.copy2(template_path, output_path)
        try:
            import win32com.client
            import pythoncom
        except ImportError:
            raise RuntimeError("pywin32 is required for COM Excel automation")

        com_initialized = False
        xl = None
        wb = None
        owns_excel = False
        original_settings = {}
        try:
            pythoncom.CoInitialize()
            com_initialized = True
            try:
                xl = win32com.client.GetActiveObject("Excel.Application")
            except pythoncom.com_error:
                xl = win32com.client.Dispatch("Excel.Application")
                owns_excel = True
            if not owns_excel:
                for setting_name in (
                    "DisplayAlerts",
                    "EnableEvents",
                    "Visible",
                    "AskToUpdateLinks",
                    "Calculation",
                    "ScreenUpdating",
                ):
                    original_settings[setting_name] = getattr(xl, setting_name)
            xl.DisplayAlerts = False
            xl.EnableEvents = False
            xl.Visible = False
            xl.AskToUpdateLinks = False
            try:
                xl.Calculation = -4135  # xlManual
            except pythoncom.com_error:
                logger.warning("Excel Calculation property could not be set")
            xl.ScreenUpdating = False

            wb = xl.Workbooks.Open(str(output_path.resolve()), UpdateLinks=False)
            ws_data = wb.Sheets("data") if "data" in [s.Name for s in wb.Sheets] else wb.Sheets[0]

            ws_data.Cells.ClearContents()

            for i, row_data in enumerate(rows, start=2):
                for j, val in enumerate(row_data, start=1):
                    ws_data.Cells(i, j).Value = val

            wb.Save()

            wb.Application.Run("RunBatch")
        except pythoncom.com_error as exc:
            raise RuntimeError(
                "Excel COM operation failed. Start the backend from the logged-in interactive "
                f"Windows user session with Excel available. Original error: {exc}"
            ) from exc
        finally:
            if xl is not None:
                if wb is not None:
                    try:
                        wb.Close(False)
                    except pythoncom.com_error:
                        pass
                if not owns_excel:
                    for setting_name, original_value in original_settings.items():
                        try:
                            setattr(xl, setting_name, original_value)
                        except pythoncom.com_error:
                            pass
                if owns_excel:
                    try:
                        xl.Quit()
                    except pythoncom.com_error:
                        logger.warning("Excel could not be quit cleanly")
            if com_initialized:
                pythoncom.CoUninitialize()

    History.objects.filter(history_id__in=target_history_ids).update(
        is_sheet_issued=True, updated_by=user, updated_at=timezone.now()
    )

    return {
        "issued_count": len(rows) + len(class9_successful_ids),
        "class9_printed": len(class9_successful_ids),
        "class9_failed": len(class9_warnings),
        "warnings": class9_warnings,
        "date": str(target_date) if target_date else "all",
        "excel_path": str(output_path) if output_path else "",
    }


def generate_daily_report(target_date, user=None):
    template_path = Path(settings.DAILY_REPORT_TEMPLATE)
    if not template_path.exists():
        raise FileNotFoundError("日報テンプレートが保存場所に存在しません。")

    output_dir = Path(settings.DAILY_REPORT_OUTPUT_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)
    owner_suffix = f"u{user.pk}" if user is not None else "legacy"
    output_path = output_dir / f"{target_date}_{owner_suffix}.xlsm"
    return _write_daily_report(template_path, output_path, target_date, user=user)


def issue_daily_report(target_date, user=None):
    template_path = Path(settings.DAILY_REPORT_TEMPLATE)
    if not template_path.exists():
        raise FileNotFoundError("日報テンプレートが保存場所に存在しません。")

    output_dir = Path(settings.DAILY_REPORT_OUTPUT_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)
    date_str = target_date.strftime("%Y%m%d")
    owner_suffix = f"u{user.pk}" if user is not None else "legacy"
    output_path = output_dir / f"{date_str}_{owner_suffix}.xlsx"
    result = _write_daily_report(template_path, output_path, target_date, keep_vba=False, user=user)

    import win32com.client
    import pythoncom

    pythoncom.CoInitialize()
    xl = None
    try:
        xl = win32com.client.DispatchEx("Excel.Application")
        xl.DisplayAlerts = False
        xl.EnableEvents = False
        xl.Visible = False
        wb = xl.Workbooks.Open(str(output_path.resolve()))
        ws = wb.Sheets("日報") if "日報" in [s.Name for s in wb.Sheets] else wb.Sheets[0]
        ws.PageSetup.BlackAndWhite = True
        ws.PrintOut()
        wb.Close(False)
    finally:
        if xl is not None:
            try:
                xl.EnableEvents = True
            except pythoncom.com_error:
                pass
            try:
                xl.Quit()
            except pythoncom.com_error:
                pass
        pythoncom.CoUninitialize()

    return result


def _write_daily_report(template_path, output_path, target_date, keep_vba=True, user=None):
    workbook = load_workbook(template_path, keep_vba=keep_vba)
    sheet = workbook["日報"] if "日報" in workbook.sheetnames else workbook.active

    slot_columns = {"A": "C", "B": "F", "C": "I", "D": "L"}
    category_start_rows = {
        2: 4,
        3: 4,
        8: 4,
        6: 18,
        7: 22,
        1: 26,
        9: 35,
    }
    category_45_counts = {slot: 0 for slot in CHECK_SLOTS}

    histories = list(
        History.objects.filter(date=target_date, created_by=user, deleted_at__isnull=True)
        .select_related("master")
        .order_by("time_slot", "master__code")
    )

    master_ids = list({h.master_id for h in histories})
    master_classes = {}
    for mc in MasterClass.objects.filter(master_id__in=master_ids).select_related("class_master"):
        master_classes[mc.master_id] = mc.class_master.class_no if mc.class_master else None

    def _category_of(history):
        return history.class_override or master_classes.get(history.master_id) or history.master.category

    # Classes sharing a start row (2, 3, 8 all start at row 4) must share a
    # single contiguous offset, otherwise they overwrite each other in the same
    # cells. Place them in spec order so class 2 (半自動機) fills from row 4.
    top_block_priority = {2: 0, 3: 1, 8: 2}
    histories.sort(
        key=lambda h: (
            h.time_slot,
            top_block_priority.get(_category_of(h), 9),
            h.master.code,
        )
    )

    row_offsets = {slot: {} for slot in CHECK_SLOTS}

    for history in histories:
        slot = history.time_slot
        category = _category_of(history)
        column = slot_columns[slot]

        if category in (4, 5):
            category_45_counts[slot] += 1
            continue

        start_row = category_start_rows.get(category)
        if start_row is None:
            continue

        offset = row_offsets[slot].get(start_row, 0)
        row = start_row + offset
        sheet[f"{column}{row}"] = history.master.name
        row_offsets[slot][start_row] = offset + 1

    for slot, count in category_45_counts.items():
        sheet[f"{slot_columns[slot]}38"] = count

    workbook.save(output_path)
    return {
        "date": str(target_date),
        "excel_path": str(output_path),
    }


def print_inspection_file(target_id, user=None):
    import os as os_mod

    target = InspectionTarget.objects.select_related("master").get(id=target_id, session__owner_user=user)
    if not target.master:
        raise ValueError("検査対象にマスターが登録されていません")

    insp_file = resolve_inspection_file(target.master, target.class_override)
    if insp_file is None and target.class_override is None:
        insp_file = resolve_unambiguous_inspection_file(target.master)
    if not insp_file:
        raise FileNotFoundError(f"検査書ファイルが見つかりません（コード: {target.normalized_code}）")

    file_path = insp_file["file_path"] if isinstance(insp_file, dict) else insp_file.file_path
    if not os_mod.path.exists(file_path):
        raise FileNotFoundError("検査書ファイルが保存場所に存在しません。")

    import tempfile
    import pythoncom
    import win32com.client
    import win32api

    ext = os_mod.path.splitext(file_path)[1].lower()

    pythoncom.CoInitialize()
    try:
        if ext in (".xls", ".xlsx", ".xlsm"):
            xl = None
            try:
                xl = win32com.client.DispatchEx("Excel.Application")
                xl.DisplayAlerts = False
                xl.Visible = False
                wb = xl.Workbooks.Open(file_path)
                for ws in wb.Sheets:
                    ws.PageSetup.BlackAndWhite = True
                wb.PrintOut()
                wb.Close(False)
            finally:
                if xl:
                    xl.Quit()
        elif ext == ".pdf":
            import fitz
            doc = fitz.open(file_path)
            new_doc = fitz.open()
            tmp_path = None
            try:
                for page in doc:
                    rect = page.rect
                    pix = page.get_pixmap(dpi=200, colorspace=fitz.csGRAY)
                    new_page = new_doc.new_page(
                        width=rect.width, height=rect.height
                    )
                    new_page.insert_image(rect, stream=pix.tobytes("png"))
                tmp_path = tempfile.mktemp(suffix=".pdf")
                new_doc.save(tmp_path)
                win32api.ShellExecute(0, "print", tmp_path, None, ".", 0)
            finally:
                doc.close()
                new_doc.close()
                if tmp_path and os_mod.path.exists(tmp_path):
                    try:
                        os_mod.unlink(tmp_path)
                    except OSError:
                        pass
        else:
            win32api.ShellExecute(0, "print", file_path, None, ".", 0)
    finally:
        pythoncom.CoUninitialize()


@transaction.atomic
def sync_all_master_classes():
    from .models import Master, Machine
    masters = Master.objects.all()
    updated = 0
    for master in masters:
        ac = get_assignment_class_for_master(master)
        if ac is not None and ac in (1, 2):
            cm = ClassMaster.objects.filter(class_no=ac).first()
            if cm:
                MasterClass.objects.update_or_create(
                    master=master,
                    class_master=cm,
                    defaults={"class_master": cm},
                )
                updated += 1
                continue
        if assigned_to_machine_class(master, 3):
            cm = ClassMaster.objects.filter(class_no=3).first()
            if cm:
                MasterClass.objects.update_or_create(
                    master=master,
                    class_master=cm,
                    defaults={"class_master": cm},
                )
                updated += 1
                continue
        n1 = (master.node_type_1 or "").strip()
        dept = (master.department or "").strip()
        if n1 == "プレス":
            cm = ClassMaster.objects.filter(class_no=4).first()
            if cm:
                MasterClass.objects.update_or_create(
                    master=master,
                    class_master=cm,
                    defaults={"class_master": cm},
                )
                updated += 1
                continue
        if n1 == "加工" and (dept == "製造管理部" or dept == "生残技術部"):
            cm = ClassMaster.objects.filter(class_no=5).first()
            if cm:
                MasterClass.objects.update_or_create(
                    master=master,
                    class_master=cm,
                    defaults={"class_master": cm},
                )
                updated += 1
                continue
        MasterClass.objects.filter(master=master).delete()
        updated += 1
    return updated


@transaction.atomic
def sync_targets_inspection_sheet_required():
    from .models import InspectionTarget
    targets = InspectionTarget.objects.filter(master__isnull=False).select_related("master")
    updated_count = 0
    for target in targets:
        required = inspection_sheet_required_for_class(target.class_override) if target.class_override else inspection_sheet_required(target.master)
        if target.requires_inspection_sheet != required:
            target.requires_inspection_sheet = required
            if required and target.issue_status == InspectionTarget.IssueStatus.NOT_REQUIRED:
                target.issue_status = InspectionTarget.IssueStatus.PENDING
            elif not required:
                target.issue_status = InspectionTarget.IssueStatus.NOT_REQUIRED
            target.save(update_fields=["requires_inspection_sheet", "issue_status", "updated_at"])
            updated_count += 1
    return updated_count
